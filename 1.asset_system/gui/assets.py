import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import sys
import os

# 添加项目根目录到路径（兼容开发和打包环境）
def get_project_root():
    """获取项目根目录"""
    if getattr(sys, 'frozen', False):
        if hasattr(sys, '_MEIPASS'):
            return sys._MEIPASS
        else:
            return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

project_root = get_project_root()
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from utils.utils import Utils
import os, shutil
from datetime import datetime


# 为Text创建placeholder支持
class PlaceholderText(tk.Text):
    def __init__(self, parent, placeholder, **kwargs):
        super().__init__(parent, **kwargs)
        self.placeholder = placeholder
        self.placeholder_color = 'gray'
        self.default_fg_color = self.cget('foreground')

        self.bind("<FocusIn>", self.clear_placeholder)
        self.bind("<FocusOut>", self.set_placeholder)
        self.set_placeholder()

    def set_placeholder(self, event=None):
        if not self.get("1.0", "end-1c"):
            self.insert("1.0", self.placeholder)
            self.config(foreground=self.placeholder_color)

    def clear_placeholder(self, event=None):
        if self.get("1.0", "end-1c") == self.placeholder:
            self.delete("1.0", tk.END)
            self.config(foreground=self.default_fg_color)

    def get_value(self):
        value = self.get("1.0", "end-1c")
        if value == self.placeholder:
            return ""
        return value

class AssetManagement:
    def __init__(self, parent, db, auth):
        self.parent = parent
        self.db = db
        self.auth = auth
        self.utils = Utils()
        self.image_references = []

        # 添加搜索状态记录
        self.is_searching = False
        self.last_search_keyword = ""

        # 资产字段索引常量
        self.ASSET_FIELD_INDICES = {
            'id': 0,
            'name': 1,
            'category': 2,
            'management_type': 3,
            'asset_number': 4,
            'quantity': 5,
            'model': 6,
            'purchase_date': 7,
            'market_value': 8,
            'responsible_person': 9,
            'location': 10,
            'status': 11,
            'lease_start_date': 12,
            'lease_end_date': 13,
            'lease_reminder_days': 14,
            'tenant_name': 15,
            'tenant_contact': 16,
            'tenant_nature': 17,
            'tenant_purpose': 18,
            'rent_amount': 19,
            'rent_payment_method': 20,
            'bidding_situation': 21,
            'certificate_status': 22,  # 产证情况
            'property_unit': 23,  # 产权单位
            'building_area': 24,  # 建筑面积
            'trusteeship_contract_type': 25,  # 合同类型
            'trusteeship_contract_amount': 26,  # 合同金额
            'trusteeship_counterparty': 27,  # 合同相对方
            'trusteeship_contract_number': 28,  # 合同编号
            'trusteeship_start_date': 29,  # 合同开始日期
            'trusteeship_end_date': 30,  # 合同结束日期
            'trusteeship_sign_date': 31,  # 签署日期
            'trusteeship_is_archived': 32,  # 是否归档
            'image_path1': 33,
            'image_path2': 34,
            'image_path3': 35,
            'notes': 36,
            'created_by': 37,
            'created_at': 38,
            'department_id': 39
        }

        # 获取当前用户信息
        self.current_user = auth.get_current_user()
        self.current_dept_id = self.current_user.get('department_id')
        self.current_dept_name = self.get_department_name(self.current_dept_id)

        # 获取部门列表和ID映射
        self.department_ids = self.load_department_mapping()

        # 添加样式配置
        style = ttk.Style()
        style.configure('FixedWidth.TCombobox', width=20)

        self.clear_content()
        self.create_asset_interface()
        self.load_assets()

    def get_department_name(self, dept_id):
        """根据部门ID获取部门名称"""
        if not dept_id:
            return "未分配部门"

        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT name FROM departments WHERE id=?", (dept_id,))
        result = c.fetchone()
        conn.close()

        return result[0] if result else "未知部门"

    def load_department_mapping(self):
        """加载部门名称到ID的映射"""
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT id, name FROM departments")
        departments = c.fetchall()
        conn.close()

        return {dept[1]: dept[0] for dept in departments}

    def clear_content(self):
        """清除内容区域"""
        for widget in self.parent.winfo_children():
            widget.destroy()

    def create_asset_interface(self):
        """创建资产管理界面"""
        # 顶部工具栏
        toolbar = ttk.Frame(self.parent, padding=10)
        toolbar.pack(fill="x")

        # 根据权限显示按钮
        if self.auth.has_permission("assets", "add"):
            ttk.Button(toolbar, text="添加资产", command=self.add_asset).pack(side="left", padx=5)

        if self.auth.has_permission("assets", "edit"):
            ttk.Button(toolbar, text="编辑资产", command=self.edit_asset).pack(side="left", padx=5)

        if self.auth.has_permission("assets", "delete"):
            ttk.Button(toolbar, text="删除资产", command=self.delete_asset).pack(side="left", padx=5)

        if self.auth.has_permission("assets", "export"):
            ttk.Button(toolbar, text="导出Excel", command=self.export_to_excel).pack(side="left", padx=5)

        if self.auth.has_permission("assets", "import"):
            ttk.Button(toolbar, text="导入Excel", command=self.import_from_excel).pack(side="left", padx=5)

        ttk.Button(toolbar, text="刷新", command=self.load_assets).pack(side="left", padx=5)

        # 租赁提醒按钮
        if self.auth.has_permission("assets", "view"):
            ttk.Button(toolbar, text="租赁提醒", command=self.show_lease_reminders).pack(side="left", padx=5)

        # 托管提醒按钮
        if self.auth.has_permission("assets", "view"):
            ttk.Button(toolbar, text="托管提醒", command=self.show_trusteeship_reminders).pack(side="left", padx=5)

        # 创建红色按钮，使用tk.Button以便设置颜色
        clear_all_btn = tk.Button(
            toolbar,
            text="清空所有资产",
            command=self.clear_all_assets,
            bg="#ff4d4d",  # 红色背景
            fg="white",  # 白色文字
            font=('微软雅黑', 10, 'bold'),
            padx=10,
            pady=5,
            relief="raised",
            bd=1,
            cursor="hand2"
        )
        clear_all_btn.pack(side="left", padx=10)

        # 添加鼠标悬停效果
        def on_enter(e):
            clear_all_btn.config(bg="#ff3333")  # 悬停时更深的红色

        def on_leave(e):
            clear_all_btn.config(bg="#ff4d4d")  # 恢复原红色

        clear_all_btn.bind("<Enter>", on_enter)
        clear_all_btn.bind("<Leave>", on_leave)

        # 搜索和筛选区域
        search_frame = ttk.Frame(self.parent, padding=10)
        search_frame.pack(fill="x")

        ttk.Label(search_frame, text="搜索:").pack(side="left")
        self.search_entry = ttk.Entry(search_frame, width=40)
        self.search_entry.pack(side="left", padx=5)
        ttk.Button(search_frame, text="搜索", command=self.search_assets).pack(side="left", padx=5)
        ttk.Button(search_frame, text="重置", command=self.reset_asset_search).pack(side="left", padx=5)

        # 数据显示表格
        tree_frame = ttk.Frame(self.parent)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # 表格列
        self.tree = ttk.Treeview(tree_frame,
                                 columns=("ID", "名称", "管理方式", "资产分类", "编号", "数量",
                                          "购置日期", "市场价值", "责任人", "地址", "状态", "租赁状态", "录入人",
                                          "所属部门"),
                                 show="headings")

        # 设置列宽和标题
        columns = {
            "ID": {"width": 50, "anchor": "center"},
            "名称": {"width": 180, "anchor": "center"},
            "管理方式": {"width": 70, "anchor": "center"},
            "资产分类": {"width": 70, "anchor": "center"},
            "编号": {"width": 70, "anchor": "center"},
            "数量": {"width": 50, "anchor": "center"},
            "购置日期": {"width": 100, "anchor": "center"},
            "市场价值": {"width": 80, "anchor": "center"},
            "责任人": {"width": 70, "anchor": "center"},
            "地址": {"width": 160, "anchor": "center"},
            "状态": {"width": 70, "anchor": "center"},
            "租赁状态": {"width": 70, "anchor": "center"},
            "录入人": {"width": 80, "anchor": "center"},
            "所属部门": {"width": 100, "anchor": "center"}
        }

        for col, settings in columns.items():
            self.tree.column(col, **settings)
            self.tree.heading(col, text=col)

        # 绑定双击事件
        self.tree.bind("<Double-1>", self.show_asset_detail)

        # 滚动条
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 状态栏
        status_bar = ttk.Frame(self.parent, height=25)
        status_bar.pack(fill="x", side="bottom")
        self.status_var = tk.StringVar()
        user = self.auth.get_current_user()
        role_display = "管理员" if user['role'] == 'admin' else '科长' if user['role'] == 'section_chief' else '员工'
        self.status_var.set(
            f"欢迎, {user['full_name'] or user['username']} | 角色: {role_display} | 部门: {self.current_dept_name or '未分配'}")
        ttk.Label(status_bar, textvariable=self.status_var, relief="sunken", padding=2).pack(fill="x")

    def clear_all_assets(self):
        """清空所有资产 - 用于重新导入数据"""

        # 获取资产数量
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM assets")
        total_assets = c.fetchone()[0]
        conn.close()

        if total_assets == 0:
            messagebox.showinfo("提示", "当前没有资产数据")
            return

        # 直接询问确认
        if messagebox.askyesno("清空所有资产",
                               f"确定要清空所有 {total_assets} 条资产记录吗？\n（用于重新导入新数据）"):
            try:
                conn = self.db.get_connection()
                c = conn.cursor()

                # 删除所有资产
                c.execute("DELETE FROM assets")
                # 重置ID
                c.execute("DELETE FROM sqlite_sequence WHERE name='assets'")

                conn.commit()
                conn.close()

                # 刷新显示
                self.load_assets()

                messagebox.showinfo("成功", f"已清空 {total_assets} 条资产记录\n可以导入新数据了")

            except Exception as e:
                messagebox.showerror("错误", f"清空失败: {str(e)}")

    def reset_asset_search(self):
        """重置资产搜索"""
        self.search_entry.delete(0, tk.END)
        self.is_searching = False
        self.last_search_keyword = ""
        self.load_assets()

    # 定义按钮调用的方法
    def add_asset(self):
        """添加资产对话框"""
        # 检查用户是否有部门（员工必须有所属部门才能添加）
        current_user = self.auth.get_current_user()
        if current_user['role'] in ['section_chief', 'staff'] and not current_user.get('department_id'):
            messagebox.showerror("错误", "您的账号未分配部门，无法添加资产")
            return

        self.asset_dialog("添加资产")

    def edit_asset(self):
        """编辑资产对话框"""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一条记录")
            return

        # 获取选中资产的部门ID
        asset_id = self.tree.item(selected_item)["values"][0]
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT department_id FROM assets WHERE id=?", (asset_id,))
        result = c.fetchone()
        conn.close()

        asset_dept_id = result[0] if result else None

        # 修改权限检查逻辑
        if not self.auth.can_edit_asset(asset_dept_id):
            # 获取资产部门名称用于提示
            conn = self.db.get_connection()
            c = conn.cursor()
            c.execute("SELECT d.name FROM assets a LEFT JOIN departments d ON a.department_id=d.id WHERE a.id=?",
                      (asset_id,))
            dept_result = c.fetchone()
            conn.close()

            asset_dept_name = dept_result[0] if dept_result else "未知部门"
            user_dept_name = self.get_department_name(self.auth.get_current_user().get('department_id'))

            messagebox.showerror("权限不足", f"您只能编辑本部门({user_dept_name})的资产\n此资产属于：{asset_dept_name}")
            return

        item_data = self.tree.item(selected_item)["values"]
        self.asset_dialog("编辑资产", item_data)

    def delete_asset(self):
        """删除资产"""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一条记录")
            return

        # 从标签中获取真实的数据库ID
        tags = self.tree.item(selected_item)["tags"]
        if not tags:
            messagebox.showerror("错误", "无法获取资产ID")
            return

        asset_id = int(tags[0].split("_")[1])

        # 获取选中资产的部门ID
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT department_id FROM assets WHERE id=?", (asset_id,))
        result = c.fetchone()
        conn.close()

        asset_dept_id = result[0] if result else None

        # 修改权限检查逻辑
        if not self.auth.can_delete_asset(asset_dept_id):
            # 获取资产部门名称用于提示
            conn = self.db.get_connection()
            c = conn.cursor()
            c.execute("SELECT d.name FROM assets a LEFT JOIN departments d ON a.department_id=d.id WHERE a.id=?",
                      (asset_id,))
            dept_result = c.fetchone()
            conn.close()

            asset_dept_name = dept_result[0] if dept_result else "未知部门"
            user_dept_name = self.get_department_name(self.auth.get_current_user().get('department_id'))

            messagebox.showerror("权限不足", f"您只能删除本部门({user_dept_name})的资产\n此资产属于：{asset_dept_name}")
            return

        if messagebox.askyesno("确认", "确定要删除这条记录吗？"):
            # 再次检查权限
            if not self.auth.can_delete_asset(asset_dept_id):
                messagebox.showerror("权限不足", "您没有删除此资产的权限")
                return

            conn = self.db.get_connection()
            c = conn.cursor()

            try:
                # 删除记录
                c.execute("DELETE FROM assets WHERE id=?", (asset_id,))

                # 重新编号
                c.execute("SELECT id FROM assets WHERE id > ? ORDER BY id", (asset_id,))
                higher_ids = [row[0] for row in c.fetchall()]

                # 依次递减ID
                for old_id in higher_ids:
                    new_id = old_id - 1
                    c.execute("UPDATE assets SET id = ? WHERE id = ?", (new_id, old_id))

                # 更新自增计数器
                c.execute("SELECT MAX(id) FROM assets")
                max_id = c.fetchone()[0]
                if max_id is None:
                    c.execute("DELETE FROM sqlite_sequence WHERE name='assets'")
                else:
                    c.execute("UPDATE sqlite_sequence SET seq = ? WHERE name='assets'", (max_id,))

                conn.commit()
                if self.is_searching and self.last_search_keyword:
                    self.search_assets()
                else:
                    self.load_assets()

                messagebox.showinfo("成功", "记录已删除")
            except Exception as e:
                conn.rollback()
                messagebox.showerror("错误", f"删除失败: {e}")
            finally:
                conn.close()

    def get_lease_status(self, management_type, lease_end_date):
        """获取租赁状态"""
        if management_type != "租赁管理" or not lease_end_date:
            return "-"

        try:
            # 检查日期是否是有效的字符串
            if isinstance(lease_end_date, str) and lease_end_date.strip():
                # 清理日期字符串，移除时间部分
                date_str = lease_end_date.strip()

                # 如果包含空格，取空格前的部分
                if ' ' in date_str:
                    date_str = date_str.split(' ')[0]

                # 确保日期字符串不为空
                if not date_str:
                    return "-"

                # 尝试解析日期
                end_date = datetime.strptime(date_str, "%Y-%m-%d")
                today = datetime.now()

                if end_date < today:
                    return "已过期"
                elif (end_date - today).days <= 30:
                    return "即将到期"
                else:
                    return "正常"
            else:
                return "-"
        except Exception as e:
            print(f"计算租赁状态出错: {e}, 日期值: {lease_end_date}, 类型: {type(lease_end_date)}")
            return "未知"


    def asset_dialog(self, title, data=None):
        """资产添加/编辑对话框"""
        dialog = tk.Toplevel(self.parent)
        dialog.title(title)
        dialog.grab_set()
        dialog.geometry("400x400")

        # 居中显示
        self.utils.center_window(dialog)

        # 使用Notebook实现标签页
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # 基本信息标签页 - 始终显示
        basic_frame = ttk.Frame(notebook)
        notebook.add(basic_frame, text="基本信息")

        # 房屋信息标签页 - 初始隐藏，当资产分类为"房屋资产"时显示
        house_frame = ttk.Frame(notebook)
        notebook.add(house_frame, text="房屋信息")

        # 租赁信息标签页 - 默认隐藏，当管理方式为"租赁管理"时显示
        lease_frame = ttk.Frame(notebook)
        notebook.add(lease_frame, text="租赁信息")

        # 托管信息 - 根据管理方式显示/隐藏
        trusteeship_frame = ttk.Frame(notebook)
        notebook.add(trusteeship_frame, text="托管信息")

        # 创建托管信息框架的网格布局
        trusteeship_grid_frame = ttk.Frame(trusteeship_frame)
        trusteeship_grid_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 创建一个使用grid布局的容器
        lease_grid_frame = ttk.Frame(lease_frame)
        lease_grid_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 其他信息标签页 - 始终显示
        other_frame = ttk.Frame(notebook)
        notebook.add(other_frame, text="其他信息")

        # 获取部门列表用于下拉框
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT id, name FROM departments ORDER BY name")
        departments = c.fetchall()
        conn.close()

        # 获取当前用户信息
        current_user = self.auth.get_current_user()

        # 根据用户角色设置部门选择
        if current_user['role'] in ['section_chief', 'staff']:
            # 科长和员工只能选择自己部门
            user_dept_id = current_user.get('department_id')
            if user_dept_id:
                # 只获取用户所在部门
                conn = self.db.get_connection()
                c = conn.cursor()
                c.execute("SELECT id, name FROM departments WHERE id=?", (user_dept_id,))
                departments = c.fetchall()
                conn.close()

            department_choices = [dept[1] for dept in departments] if departments else ["未分配部门"]
        else:
            # 管理员可以选择所有部门
            department_choices = ["未选择"] + [dept[1] for dept in departments]

        department_ids = {dept[1]: dept[0] for dept in departments}

        entries = {}


        # 自定义Entry类，支持placeholder
        class PlaceholderEntry(ttk.Entry):
            def __init__(self, parent, placeholder, **kwargs):
                super().__init__(parent, **kwargs)
                self.placeholder = placeholder
                self.placeholder_color = 'gray'
                self.default_fg_color = self.cget('foreground')

                self.bind("<FocusIn>", self.clear_placeholder)
                self.bind("<FocusOut>", self.set_placeholder)
                self.set_placeholder()

            def set_placeholder(self, event=None):
                if not self.get():
                    self.insert(0, self.placeholder)
                    self.config(foreground=self.placeholder_color)

            def clear_placeholder(self, event=None):
                if self.get() == self.placeholder:
                    self.delete(0, tk.END)
                    self.config(foreground=self.default_fg_color)

            def get_value(self):
                value = self.get()
                if value == self.placeholder:
                    return ""
                return value

        # 基本信息字段
        basic_fields = [
            ("资产名称", "name", 0, "请输入资产名称"),
            ("管理方式", "management_type", 1, None),
            ("资产分类", "category", 2, None),
            ("资产编号", "asset_number", 3, "如：GY2025"),
            ("资产数量", "quantity", 4, "请输入整数"),
            ("购置日期", "purchase_date", 5, "YYYY-MM-DD"),
            ("市场价值", "market_value", 6, "单位：元，可含小数"),
            ("所属部门", "department", 7, None)
        ]

        for i, (label, field, row, placeholder) in enumerate(basic_fields):
            ttk.Label(basic_frame, text=label + ":").grid(row=row, column=0, padx=5, pady=5, sticky="e")

            if field == "management_type":
                # 添加"托管管理"选项
                entry = ttk.Combobox(basic_frame,
                                     values=["自主管理", "租赁管理", "托管管理"],
                                     state="readonly",
                                     style='FixedWidth.TCombobox')
                entry.set("自主管理")  # 默认值
                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
            elif field == "category":
                entry = ttk.Combobox(basic_frame,
                                     values=["房屋资产", "办公资产", "其他资产"],
                                     state="readonly",
                                     style='FixedWidth.TCombobox')
                entry.set("房屋资产")
                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
            elif field == "department":
                entry = ttk.Combobox(basic_frame, values=department_choices, state="readonly",
                                     style='FixedWidth.TCombobox')

                # 获取当前用户信息
                current_user = self.auth.get_current_user()

                # 根据用户角色设置部门字段
                if current_user['role'] in ['section_chief', 'staff']:
                    # 科长和员工只能选择本部门，且为只读
                    if department_choices:
                        entry.set(department_choices[0])
                    else:
                        entry.set("未分配部门")
                    entry.config(state="readonly")  # 设置为只读，不能选择其他部门
                else:
                    # 管理员可以选择任何部门
                    if data and len(data) > 39:  # 编辑模式，需要设置部门
                        # 从数据中获取部门ID并查找名称
                        asset_dept_id = data[39]
                        if asset_dept_id:
                            conn = self.db.get_connection()
                            c = conn.cursor()
                            c.execute("SELECT name FROM departments WHERE id=?", (asset_dept_id,))
                            dept_result = c.fetchone()
                            conn.close()
                            if dept_result:
                                entry.set(dept_result[0])
                            else:
                                entry.set("未选择")
                        else:
                            entry.set("未选择")
                    else:
                        # 添加模式，默认未选择
                        entry.set("未选择")

                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")

            else:
                entry = PlaceholderEntry(basic_frame, placeholder, width=30)
                entry.grid(row=row, column=1, padx=5, pady=5)

            entries[field] = entry

        # 房屋信息字段
        house_fields = [
            ("产证情况", "certificate_status", 0, "如房产证、土地证信息"),
            ("产权单位", "property_unit", 1, "请输入产权单位名称"),
            ("建筑面积", "building_area", 2, "单位：平方米")
        ]

        for i, (label, field, row, placeholder) in enumerate(house_fields):
            ttk.Label(house_frame, text=label + ":").grid(row=row, column=0, padx=5, pady=5, sticky="e")

            if field == "certificate_status":
                # 产证情况使用多行文本框
                text_frame = ttk.Frame(house_frame)
                text_frame.grid(row=row, column=1, padx=5, pady=5, sticky="w")

                entry = PlaceholderText(text_frame, placeholder, width=30, height=3)  # 3行高度
                entry.pack(side="left")

                scrollbar = ttk.Scrollbar(text_frame, command=entry.yview)
                scrollbar.pack(side="right", fill="y")
                entry.config(yscrollcommand=scrollbar.set)
            elif field == "building_area":
                # 建筑面积添加单位提示
                unit_frame = ttk.Frame(house_frame)
                unit_frame.grid(row=row, column=1, padx=5, pady=5, sticky="w")

                entry = PlaceholderEntry(unit_frame, placeholder, width=25)
                entry.pack(side="left")

                # 添加单位标签
                ttk.Label(unit_frame, text="平方米", font=('微软雅黑', 9), foreground="black").pack(side="left",
                                                                                                   padx=(5, 0))
            else:
                entry = PlaceholderEntry(house_frame, placeholder, width=30)
                entry.grid(row=row, column=1, padx=5, pady=5)

            entries[field] = entry

        # 租赁信息字段
        lease_fields = [
            ("租赁开始日期", "lease_start_date", 0, "YYYY-MM-DD"),
            ("租赁结束日期", "lease_end_date", 1, "YYYY-MM-DD"),
            ("提前提醒天数", "lease_reminder_days", 2, "1-365天"),
            ("承租方名称", "tenant_name", 3, "输入承租方全称"),
            ("承租方联系方式", "tenant_contact", 4, "输入电话或手机号"),
            ("承租方性质", "tenant_nature", 5, None),
            ("承租方用途", "tenant_purpose", 6, "如经营XX店铺等"),
            ("租金金额", "rent_amount", 7, "单位：元，可含小数"),
            ("租金交付方式", "rent_payment_method", 8, "如：半年交/全年交"),
            ("公开招拍租情况", "bidding_situation", 9, "如：公开拍租/公开拍租中")
        ]

        # 创建所有租赁字段
        for label, field, row, placeholder in lease_fields:
            # 标签
            ttk.Label(lease_grid_frame, text=label + ":").grid(
                row=row, column=0, padx=5, pady=5, sticky="e")

            # 输入控件
            if field == "lease_reminder_days":
                entry = ttk.Spinbox(lease_grid_frame, from_=1, to=365, width=27)
                entry.set("30")
            elif field == "tenant_nature":
                entry = ttk.Combobox(lease_grid_frame, values=["个人", "公司"],
                                     state="readonly", width=27)
                entry.set("个人")
            elif field == "rent_amount":
                # 创建带单位的输入框
                unit_frame = ttk.Frame(lease_grid_frame)
                unit_frame.grid(row=row, column=1, padx=5, pady=5, sticky="w")

                entry = PlaceholderEntry(unit_frame, placeholder, width=25)
                entry.pack(side="left")

                ttk.Label(unit_frame, text="元", font=('微软雅黑', 9)).pack(side="left", padx=(5, 0))
            else:
                entry = PlaceholderEntry(lease_grid_frame, placeholder, width=30)
                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")

            # 如果不是租金金额字段，设置grid布局
            if field != "rent_amount":
                if field == "lease_reminder_days" or field == "tenant_nature":
                    entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
                else:
                    entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")

            entries[field] = entry

        # 新增托管信息字段
        trusteeship_fields = [
            ("合同类型", "trusteeship_contract_type", 1, "请输入合同类型"),
            ("合同金额", "trusteeship_contract_amount", 2, "请输入合同金额"),
            ("合同相对方", "trusteeship_counterparty", 3, "请输入合同相对方名称"),
            ("合同编号", "trusteeship_contract_number", 4, "请输入合同编号"),
            ("合同开始日期", "trusteeship_start_date", 5, "YYYY-MM-DD"),
            ("合同结束日期", "trusteeship_end_date", 6, "YYYY-MM-DD"),
            ("签署日期", "trusteeship_sign_date", 7, "YYYY-MM-DD"),
            ("是否归档", "trusteeship_is_archived", 8, None),
        ]

        for label, field, row, placeholder in trusteeship_fields:
            ttk.Label(trusteeship_grid_frame, text=label + ":").grid(
                row=row, column=0, padx=5, pady=5, sticky="e")

            if field == "trusteeship_is_archived":
                entry = ttk.Combobox(trusteeship_grid_frame, values=["是", "否"],
                                     state="readonly", width=27)
                entry.set("否")
                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
            elif field == "trusteeship_contract_amount":
                # 修改为纯文本输入框，不带单位
                entry = PlaceholderEntry(trusteeship_grid_frame, placeholder, width=30)
                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
            elif field == "trusteeship_counterparty":  # 合同相对方使用多行文本框
                # 创建多行文本框框架
                text_frame = ttk.Frame(trusteeship_grid_frame)
                text_frame.grid(row=row, column=1, padx=5, pady=5, sticky="w")

                entry = PlaceholderText(text_frame, placeholder, width=30, height=3)
                entry.pack(side="left")

                # 添加滚动条
                scrollbar = ttk.Scrollbar(text_frame, command=entry.yview)
                scrollbar.pack(side="right", fill="y")
                entry.config(yscrollcommand=scrollbar.set)
            else:
                entry = PlaceholderEntry(trusteeship_grid_frame, placeholder, width=30)
                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")

            entries[field] = entry

        # 其他信息字段
        other_fields = [
            ("资产责任人", "responsible_person", 0, "输入责任人姓名"),
            ("资产地址", "location", 1, "详细地址，如：XX市XX区XX路XX号"),
            ("经度", "longitude", 2, "如：116.397428"),
            ("纬度", "latitude", 3, "如：39.90923"),
            ("坐标类型", "coord_type", 4, None),
            ("资产状态", "status", 5, None),
            ("备注", "notes", 6, None)
        ]


        for i, (label, field, row, placeholder) in enumerate(other_fields):
            ttk.Label(other_frame, text=label + ":").grid(row=row, column=0, padx=5, pady=5, sticky="e")

            if field == "status":
                entry = ttk.Combobox(other_frame,
                                     values=["正常使用", "维修中", "报废", "闲置", "公开招租中", "已终止", "已结束"],
                                     state="readonly")
                entry.set("正常使用")
                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
            elif field in ["longitude", "latitude"]:
                # 经纬度字段，添加单位提示
                unit_frame = ttk.Frame(other_frame)
                unit_frame.grid(row=row, column=1, padx=5, pady=5, sticky="w")

                entry = PlaceholderEntry(unit_frame, placeholder, width=25)
                entry.pack(side="left")

                # 添加单位标签
                unit_text = "经度范围：-180~180" if field == "longitude" else "纬度范围：-90~90"
                ttk.Label(unit_frame, text=unit_text, font=('微软雅黑', 8), foreground="gray").pack(side="left", padx=(5, 0))
            elif field == "coord_type":
                # 坐标类型下拉框
                entry = ttk.Combobox(other_frame,
                                     values=["WGS84（GPS/国际）", "GCJ02（高德/火星）", "BD09（百度）"],
                                     state="readonly", width=25)
                entry.set("WGS84（GPS/国际）")
                entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
                # 提示
                ttk.Label(other_frame, text="* 百度坐标需选BD09",
                          font=('微软雅黑', 8), foreground="gray").grid(row=row, column=2, padx=5, sticky="w")
            elif field == "notes":
                text_frame = ttk.Frame(other_frame)
                text_frame.grid(row=row, column=1, padx=5, pady=5, sticky="w")

                entry = PlaceholderText(text_frame, "可选，最多500字", width=30, height=5)
                entry.pack(side="left")

                scrollbar = ttk.Scrollbar(text_frame, command=entry.yview)
                scrollbar.pack(side="right", fill="y")
                entry.config(yscrollcommand=scrollbar.set)
            else:
                entry = PlaceholderEntry(other_frame, placeholder, width=30)
                entry.grid(row=row, column=1, padx=5, pady=5)

            entries[field] = entry

        # 图片上传区域
        ttk.Label(other_frame, text="资产图片:").grid(row=4, column=0, padx=5, pady=5, sticky="ne")

        img_frame = ttk.Frame(other_frame)
        img_frame.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        entries["image_paths"] = []
        for i in range(3):
            frame = ttk.Frame(img_frame)
            frame.pack(fill="x", pady=2)

            btn = ttk.Button(frame, text=f"选择图片 {i + 1}",
                             command=lambda idx=i: self.select_image(entries["image_paths"][idx]))
            btn.pack(side="left")

            entry = ttk.Label(frame, text="未选择图片", width=20, foreground="gray")
            entry.pack(side="left", padx=5)

            entries["image_paths"].append(entry)

        # 标签页显示/隐藏控制函数
        def toggle_tabs():
            category = entries["category"].get()
            management_type = entries["management_type"].get()

            # 控制房屋信息标签页
            if category == "房屋资产":
                notebook.tab(1, state="normal")  # 显示房屋信息（索引1）
            else:
                notebook.tab(1, state="hidden")  # 隐藏房屋信息

            # 控制租赁信息标签页
            if management_type == "租赁管理":
                notebook.tab(2, state="normal")  # 显示租赁信息（索引2）
            else:
                notebook.tab(2, state="hidden")  # 隐藏租赁信息

            # 控制托管信息标签页
            if management_type == "托管管理":
                notebook.tab(3, state="normal")  # 显示托管信息（索引3）
            else:
                notebook.tab(3, state="hidden")  # 隐藏托管信息

            # 其他信息标签页（索引4）始终显示
            notebook.tab(4, state="normal")

        # 绑定事件
        entries["category"].bind("<<ComboboxSelected>>", lambda e: toggle_tabs())
        entries["management_type"].bind("<<ComboboxSelected>>", lambda e: toggle_tabs())

        # 初始隐藏标签页
        toggle_tabs()

        # 如果在添加模式下，检查用户是否有部门（员工必须有所属部门才能添加）
        if not data and current_user['role'] in ['section_chief', 'staff']:
            if not current_user.get('department_id'):
                messagebox.showerror("错误", "您的账号未分配部门，无法添加资产")
                dialog.destroy()
                return

        if data:
            # 获取资产ID
            asset_id = data[0]
            conn = self.db.get_connection()
            c = conn.cursor()

            # 明确指定字段顺序（包括经纬度）
            c.execute('''
                SELECT
                    id, name, category, management_type, asset_number, quantity,
                    model, purchase_date, market_value, responsible_person,
                    location, status, lease_start_date, lease_end_date,
                    lease_reminder_days, tenant_name, tenant_contact,
                    tenant_nature, tenant_purpose, rent_payment_method, bidding_situation,
                    certificate_status, property_unit, building_area,
                    rent_amount,
                    trusteeship_contract_type,
                    trusteeship_contract_amount, trusteeship_counterparty,
                    trusteeship_contract_number, trusteeship_start_date,
                    trusteeship_end_date, trusteeship_sign_date,
                    trusteeship_is_archived,
                    image_path1, image_path2, image_path3, notes,
                    created_by, created_at, department_id,
                    longitude, latitude, coord_type
                FROM assets WHERE id=?
            ''', (asset_id,))
            asset_data = c.fetchone()
            conn.close()

            if asset_data:
                # 明确的数据索引映射
                data_mapping = {
                    "name": 1,
                    "category": 2,
                    "management_type": 3,
                    "asset_number": 4,
                    "quantity": 5,
                    "purchase_date": 7,
                    "market_value": 8,
                    "responsible_person": 9,
                    "location": 10,
                    "status": 11,
                    "lease_start_date": 12,
                    "lease_end_date": 13,
                    "lease_reminder_days": 14,
                    "tenant_name": 15,
                    "tenant_contact": 16,
                    "tenant_nature": 17,
                    "tenant_purpose": 18,
                    "rent_payment_method": 19,
                    "bidding_situation": 20,
                    "certificate_status": 21,
                    "property_unit": 22,
                    "building_area": 23,
                    "rent_amount": 24,
                    "trusteeship_contract_type": 25,
                    "trusteeship_contract_amount": 26,
                    "trusteeship_counterparty": 27,
                    "trusteeship_contract_number": 28,
                    "trusteeship_start_date": 29,
                    "trusteeship_end_date": 30,
                    "trusteeship_sign_date": 31,
                    "trusteeship_is_archived": 32,
                    "image_path1": 33,
                    "image_path2": 34,
                    "image_path3": 35,
                    "notes": 36,
                    "created_by": 37,
                    "created_at": 38,
                    "department_id": 39,
                    "longitude": 40,
                    "latitude": 41,
                    "coord_type": 42
                }

                # 填充通用字段
                for field_name, data_index in data_mapping.items():
                    if data_index < len(asset_data) and asset_data[data_index] is not None:
                        value = str(asset_data[data_index])
                        entry = entries.get(field_name)

                        if entry:
                            # 清除placeholder
                            if hasattr(entry, 'clear_placeholder'):
                                entry.clear_placeholder()

                            # 坐标类型特殊处理：转换为下拉框显示文本
                            if field_name == "coord_type":
                                if value == "bd09":
                                    entry.set("BD09（百度）")
                                elif value == "gcj02":
                                    entry.set("GCJ02（高德/火星）")
                                else:
                                    entry.set("WGS84（GPS/国际）")
                            # 根据控件类型设置值
                            elif isinstance(entry, tk.Text) or isinstance(entry, PlaceholderText):
                                entry.delete("1.0", tk.END)
                                if value:
                                    entry.insert("1.0", value)
                            elif isinstance(entry, ttk.Combobox):
                                entry.set(value)
                            elif isinstance(entry, ttk.Spinbox):
                                try:
                                    entry.delete(0, tk.END)
                                    entry.insert(0, int(value))
                                except ValueError:
                                    entry.delete(0, tk.END)
                                    entry.insert(0, value)
                            else:
                                entry.delete(0, tk.END)
                                entry.insert(0, value)

                # 填充部门信息
                if len(asset_data) > 39 and asset_data[39]:
                    department_id = asset_data[39]
                    if department_id:
                        conn = self.db.get_connection()
                        c = conn.cursor()
                        c.execute("SELECT name FROM departments WHERE id=?", (department_id,))
                        dept_result = c.fetchone()
                        conn.close()

                        if dept_result:
                            entries["department"].set(dept_result[0])
                        else:
                            entries["department"].set("未选择")
                    else:
                        entries["department"].set("未选择")
                else:
                    entries["department"].set("未选择")

                # 填充图片路径
                image_indices = [33, 34, 35]  # image_path1, image_path2, image_path3
                for i, img_idx in enumerate(image_indices):
                    if i < len(entries["image_paths"]):
                        if img_idx < len(asset_data) and asset_data[img_idx]:
                            path = str(asset_data[img_idx])
                            if path and path.strip() and path not in ["未选择图片", "None", ""]:
                                # 只显示文件名
                                filename = os.path.basename(path)
                                entries["image_paths"][i].config(text=filename, foreground="black")
                                # 保存完整路径到自定义属性
                                entries["image_paths"][i].image_full_path = path
                                entries["image_paths"][i].is_image_selected = True

                # 根据分类和管理方式显示/隐藏标签页
                toggle_tabs()

        # 按钮区域
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=(0, 10), fill="x")

        # 创建一个居中的子框架
        center_frame = ttk.Frame(btn_frame)
        center_frame.pack()

        # 按钮放在居中框架里
        ttk.Button(center_frame, text="保存",
                   command=lambda: self.save_asset(dialog, entries, data[0] if data else None)).pack(
            side="left", padx=10)
        ttk.Button(center_frame, text="取消", command=dialog.destroy).pack(
            side="left", padx=10)

        # 确保对话框有合适的大小
        dialog.update_idletasks()

        # 如果对话框太小，设置最小大小
        if dialog.winfo_height() < 450:
            dialog.minsize(400, 450)

    def toggle_lease_tab(self, notebook, management_type):
        """根据管理方式显示/隐藏租赁信息标签页"""
        if management_type == "租赁管理":
            notebook.tab(1, state="normal")  # 显示租赁信息标签页
        else:
            notebook.tab(1, state="hidden")  # 隐藏租赁信息标签页

    def show_lease_reminders(self):
        """显示租赁提醒"""
        conn = self.db.get_connection()
        c = conn.cursor()

        # 获取当前用户信息
        current_user = self.auth.get_current_user()

        # 查询即将到期的租赁资产，添加部门权限限制
        query = '''SELECT a.name, a.lease_end_date, a.responsible_person, a.location, a.tenant_name
                  FROM assets a
                  WHERE a.management_type = '租赁管理'
                  AND a.lease_end_date IS NOT NULL
                  AND date(a.lease_end_date) <= date('now', '+' || COALESCE(a.lease_reminder_days, 30) || ' days')'''

        params = []

        # 部门权限控制
        if current_user['role'] in ['section_chief', 'staff']:
            query += " AND a.department_id = ?"
            params.append(current_user.get('department_id'))

        query += " ORDER BY a.lease_end_date"

        c.execute(query, params)
        reminders = c.fetchall()
        conn.close()

        if not reminders:
            messagebox.showinfo("租赁提醒", "没有即将到期的租赁资产")
            return

        # 创建提醒对话框
        reminder_dialog = tk.Toplevel(self.parent)
        reminder_dialog.title("租赁到期提醒")
        reminder_dialog.geometry("900x500")

        self.utils.center_window(reminder_dialog)

        # 创建表格显示提醒信息
        tree_frame = ttk.Frame(reminder_dialog)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        tree = ttk.Treeview(tree_frame,
                            columns=("资产名称", "到期日期", "剩余天数", "责任人", "位置", "承租方", "所属部门"),
                            show="headings")

        columns = {
            "资产名称": {"width": 200, "anchor": "center"},
            "到期日期": {"width": 80, "anchor": "center"},
            "剩余天数": {"width": 80, "anchor": "center"},
            "责任人": {"width": 80, "anchor": "center"},
            "位置": {"width": 180, "anchor": "center"},
            "承租方": {"width": 80, "anchor": "center"},
            "所属部门": {"width": 80, "anchor": "center"}
        }

        for col, settings in columns.items():
            tree.column(col, **settings)
            tree.heading(col, text=col)

        # 添加滚动条
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 填充数据
        today = datetime.now().date()
        for reminder in reminders:
            try:
                end_date = datetime.strptime(reminder[1], "%Y-%m-%d").date()
                days_left = (end_date - today).days

                status_color = ""
                if days_left < 0:
                    status_text = f"已过期{abs(days_left)}天"
                    status_color = "red"
                elif days_left <= 7:
                    status_text = f"剩余{days_left}天"
                    status_color = "orange"
                else:
                    status_text = f"剩余{days_left}天"
                    status_color = "green"

                # 获取资产所属部门
                conn = self.db.get_connection()
                c = conn.cursor()
                c.execute('''SELECT d.name FROM assets a 
                             LEFT JOIN departments d ON a.department_id = d.id 
                             WHERE a.name=? AND a.lease_end_date=?''',
                          (reminder[0], reminder[1]))
                dept_result = c.fetchone()
                conn.close()

                dept_name = dept_result[0] if dept_result else "未知部门"

                tree.insert("", "end", values=(
                    reminder[0], reminder[1], status_text,
                    reminder[2], reminder[3], reminder[4] or "未知", dept_name
                ), tags=(status_color,))

            except:
                tree.insert("", "end", values=(
                    reminder[0], reminder[1], "日期格式错误",
                    reminder[2], reminder[3], reminder[4] or "未知", "未知部门"
                ))

        # 设置颜色
        tree.tag_configure("red", foreground="red")
        tree.tag_configure("orange", foreground="orange")
        tree.tag_configure("green", foreground="green")

        # 关闭按钮
        ttk.Button(reminder_dialog, text="关闭", command=reminder_dialog.destroy).pack(pady=10)

    def show_trusteeship_reminders(self):
        """显示托管合同提醒 - 半年内到期，排除已结束/已终止的项目"""
        conn = self.db.get_connection()
        c = conn.cursor()

        # 获取当前用户信息
        current_user = self.auth.get_current_user()

        # 查询半年内到期的托管合同，排除已结束和已终止的项目
        query = '''SELECT a.name, a.trusteeship_end_date, a.responsible_person, 
                          a.status, a.trusteeship_counterparty,
                          a.department_id, d.name as department_name
                   FROM assets a
                   LEFT JOIN departments d ON a.department_id = d.id
                   WHERE a.management_type = '托管管理'
                   AND a.trusteeship_end_date IS NOT NULL
                   AND a.trusteeship_end_date != ''
                   AND a.status NOT IN ('已结束', '已终止') 
                   AND date(a.trusteeship_end_date) BETWEEN date('now') 
                   AND date('now', '+180 days') 
                   ORDER BY a.trusteeship_end_date'''

        params = []

        # 部门权限控制
        if current_user['role'] in ['section_chief', 'staff']:
            query += " AND a.department_id = ?"
            params.append(current_user.get('department_id'))

        c.execute(query, params)
        reminders = c.fetchall()
        conn.close()

        if not reminders:
            messagebox.showinfo("托管提醒", "半年内没有即将到期的托管合同\n（已结束/已终止的项目已过滤）")
            return

        # 创建提醒对话框
        reminder_dialog = tk.Toplevel(self.parent)
        reminder_dialog.title("托管合同到期提醒（半年内）")
        reminder_dialog.geometry("800x400")

        self.utils.center_window(reminder_dialog)

        # 创建表格显示提醒信息
        tree_frame = ttk.Frame(reminder_dialog)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        tree = ttk.Treeview(tree_frame,
                            columns=("资产名称", "到期日期", "剩余天数", "责任人",
                                     "状态", "合同相对方", "所属部门"),
                            show="headings")

        columns = {
            "资产名称": {"width": 150, "anchor": "center"},
            "到期日期": {"width": 100, "anchor": "center"},
            "剩余天数": {"width": 80, "anchor": "center"},
            "责任人": {"width": 80, "anchor": "center"},
            "状态": {"width": 80, "anchor": "center"},
            "合同相对方": {"width": 100, "anchor": "center"},
            "所属部门": {"width": 100, "anchor": "center"}
        }

        for col, settings in columns.items():
            tree.column(col, **settings)
            tree.heading(col, text=col)

        # 添加滚动条
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 填充数据
        today = datetime.now().date()
        for reminder in reminders:
            try:
                end_date = datetime.strptime(reminder[1], "%Y-%m-%d").date()
                days_left = (end_date - today).days

                status_color = ""
                if days_left < 0:
                    status_text = f"已过期{abs(days_left)}天"
                    status_color = "red"
                elif days_left <= 30:
                    status_text = f"剩余{days_left}天"
                    status_color = "orange"
                elif days_left <= 90:
                    status_text = f"剩余{days_left}天"
                    status_color = "#FFD700"  # 金色，90天内
                else:
                    status_text = f"剩余{days_left}天"
                    status_color = "green"

                # 获取资产状态
                asset_status = reminder[3] if reminder[3] else "正常使用"

                tree.insert("", "end", values=(
                    reminder[0],  # 资产名称
                    reminder[1],  # 到期日期
                    status_text,  # 剩余天数
                    reminder[2],  # 责任人
                    asset_status,  # 状态
                    reminder[4] or "未知",  # 合同相对方
                    reminder[6] if reminder[6] else "未知部门"  # 部门名称
                ), tags=(status_color,))

            except Exception as e:
                print(f"处理托管提醒数据出错: {e}")
                tree.insert("", "end", values=(
                    reminder[0],  # 资产名称
                    reminder[1],  # 到期日期
                    "日期格式错误",  # 剩余天数
                    reminder[2],  # 责任人
                    reminder[3] if reminder[3] else "正常使用",  # 状态
                    reminder[4] or "未知",  # 合同相对方
                    reminder[6] if reminder[6] else "未知部门"  # 部门名称
                ))

        # 设置颜色
        tree.tag_configure("red", foreground="red")
        tree.tag_configure("orange", foreground="orange")
        tree.tag_configure("green", foreground="green")
        tree.tag_configure("#FFD700", foreground="#FFD700")  # 金色配置

        # 添加颜色说明
        color_frame = ttk.Frame(reminder_dialog)
        color_frame.pack(fill="x", padx=10, pady=(0, 5))

        ttk.Label(color_frame, text="颜色说明:", font=('微软雅黑', 9, 'bold')).pack(side="left", padx=(0, 10))

        colors_info = [
            ("已过期", "red"),
            ("30天内到期", "orange"),
            ("90天内到期", "#FFD700"),
            ("90天以上", "green")
        ]

        for text, color in colors_info:
            color_label = tk.Label(color_frame, text="●", font=('微软雅黑', 12), fg=color)
            color_label.pack(side="left", padx=(10, 2))

            text_label = ttk.Label(color_frame, text=text, font=('微软雅黑', 9))
            text_label.pack(side="left", padx=(0, 10))

        # 关闭按钮
        ttk.Button(reminder_dialog, text="关闭", command=reminder_dialog.destroy).pack(pady=10)

    def export_to_excel(self):
        """导出资产到Excel"""
        try:
            if not self.auth.has_permission("assets", "export"):
                messagebox.showerror("权限不足", "您没有导出权限")
                return

            default_filename = f"资产清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialfile=default_filename
            )

            if not filepath:
                return

            conn = self.db.get_connection()
            c = conn.cursor()

            # 构建查询字段列表
            query = '''SELECT 
                        a.id, a.name, a.management_type, a.category, a.asset_number, 
                        a.quantity, a.purchase_date, a.market_value,
                        a.responsible_person, a.location, a.status,
                        a.lease_start_date, a.lease_end_date,
                        a.tenant_name, a.tenant_contact,
                        a.tenant_nature, a.tenant_purpose, a.rent_amount, a.rent_payment_method, a.bidding_situation, 
                        a.certificate_status, a.property_unit, a.building_area,
                        a.trusteeship_contract_type,
                        a.trusteeship_contract_amount, a.trusteeship_counterparty,
                        a.trusteeship_contract_number, a.trusteeship_start_date,
                        a.trusteeship_end_date, a.trusteeship_sign_date,
                        a.trusteeship_is_archived,
                        COALESCE(u.full_name, a.created_by) as display_name,
                        d.name as department_name,
                        a.notes
                       FROM assets a
                       LEFT JOIN users u ON a.created_by = u.username
                       LEFT JOIN departments d ON a.department_id = d.id'''

            params = []

            # 部门权限限制
            current_user = self.auth.get_current_user()
            if current_user['role'] in ['section_chief', 'staff']:
                query += " WHERE a.department_id = ?"
                params.append(current_user.get('department_id'))

            query += " ORDER BY a.id"

            c.execute(query, params)
            rows = c.fetchall()
            conn.close()

            if not rows:
                messagebox.showwarning("提示", "没有找到符合条件的资产记录")
                return

            # 创建Excel文件
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "资产清单"

                # 定义表头
                headers = [
                    "ID", "名称", "管理方式", "资产分类", "编号", "数量",
                    "购置日期", "市场价值", "责任人", "地址", "状态",
                    "租赁开始日期", "租赁结束日期", "承租方", "承租方联系方式",
                    "承租方性质", "承租方用途", "租金金额", "租金交付方式", "公开招拍租情况",
                    "产证情况", "产权单位", "建筑面积（平方米）",
                    "托管合同类型", "托管合同金额", "托管合同相对方",
                    "托管合同编号", "托管合同开始日期", "托管合同结束日期", "托管签署日期",
                    "托管是否归档",
                    "录入人", "所属部门", "备注"
                ]

                # 验证字段数量与查询结果匹配
                if len(rows[0]) != len(headers):
                    messagebox.showerror("错误",
                                         f"字段数量不匹配:\n查询返回 {len(rows[0])} 个字段\n表头定义 {len(headers)} 个字段\n"
                                         f"请检查SQL查询和表头定义是否一致")
                    return

                ws.append(headers)

                # 添加数据行
                for display_id, row in enumerate(rows, 1):
                    formatted_row = []

                    # 显示序号
                    formatted_row.append(display_id)

                    # 获取管理方式
                    management_type = row[2] if len(row) > 2 else ""

                    # 处理每个字段，跳过数据库ID (索引0)
                    for i in range(1, len(row)):
                        value = row[i]

                        # 托管字段检查（索引24-30）
                        is_trusteeship_field = 24 <= i <= 30

                        if is_trusteeship_field and management_type != "托管管理":
                            # 不是托管管理的资产，托管字段都为空
                            formatted_row.append("")
                        elif i == 5:  # 数量字段
                            try:
                                if value is not None:
                                    formatted_row.append(int(float(value)))
                                else:
                                    formatted_row.append("")
                            except (ValueError, TypeError):
                                formatted_row.append("")
                        elif i == 7:  # 市场价值
                            try:
                                if value is not None and value != "" and value != 0:
                                    formatted_row.append(float(value))
                                else:
                                    formatted_row.append("")
                            except:
                                formatted_row.append("")
                        elif i == 17:  # 租金金额
                            try:
                                if value is not None and value != "" and value != 0:
                                    formatted_row.append(float(value))
                                else:
                                    formatted_row.append("")
                            except:
                                formatted_row.append("")
                        elif i == 24:  # 托管合同金额
                            if management_type != "托管管理":
                                formatted_row.append("")
                            else:
                                if value is None or value == "" or value == 0:
                                    formatted_row.append("")
                                else:
                                    try:
                                        formatted_row.append(float(value))
                                    except:
                                        formatted_row.append(str(value).strip())
                        elif i == 31:  # 录入人字段 - 关键！
                            formatted_row.append(str(value).strip() if value is not None else "")
                        elif i == 32:  # 所属部门字段
                            formatted_row.append(str(value).strip() if value is not None else "")
                        elif i == 33:  # 备注字段
                            formatted_row.append(str(value).strip() if value is not None else "")
                        elif i in [6, 11, 12, 27, 28, 29]:  # 日期字段
                            if value is None or value == "" or value == "-":
                                formatted_row.append("")
                            else:
                                try:
                                    date_str = str(value).strip()
                                    if date_str:
                                        datetime.strptime(date_str, "%Y-%m-%d")
                                        formatted_row.append(date_str)
                                    else:
                                        formatted_row.append("")
                                except:
                                    formatted_row.append("")
                        else:
                            # 其他字段
                            if value is None:
                                formatted_row.append("")
                            else:
                                formatted_row.append(str(value).strip())

                    ws.append(formatted_row)

                # 修改列宽设置
                col_widths = [
                    10,  # ID显示序号
                    25,  # 名称
                    12,  # 管理方式
                    12,  # 资产分类
                    15,  # 编号
                    8,  # 数量
                    12,  # 购置日期
                    15,  # 市场价值
                    12,  # 责任人
                    25,  # 地址
                    10,  # 状态
                    15,  # 租赁开始日期
                    15,  # 租赁结束日期
                    15,  # 承租方
                    15,  # 承租方联系方式
                    12,  # 承租方性质
                    20,  # 承租方用途
                    15,  # 租金金额
                    15,  # 租金交付方式
                    20,  # 公开招拍租情况
                    20,  # 产证情况
                    20,  # 产权单位
                    20,  # 建筑面积
                    15,  # 托管合同类型
                    15,  # 托管合同金额
                    20,  # 托管合同相对方
                    15,  # 托管合同编号
                    18,  # 托管合同开始日期
                    18,  # 托管合同结束日期
                    16,  # 托管签署日期
                    12,  # 托管是否归档
                    12,  # 录入人
                    15,  # 所属部门
                    30  # 备注
                ]

                for i, width in enumerate(col_widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

                # 设置统一的行高
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    ws.row_dimensions[row[0].row].height = 30

                # 设置标题样式
                from openpyxl.styles import Font, Alignment, Border, Side

                header_font = Font(bold=True, name='微软雅黑', size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")

                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = header_alignment

                # 为日期列设置条件格式
                date_columns = ['G', 'L', 'M']  # 购置日期、租赁开始日期、租赁结束日期

                for col in date_columns:
                    for cell in ws[col][1:]:
                        if cell.value and str(cell.value).strip():
                            try:
                                datetime.strptime(str(cell.value).strip(), "%Y-%m-%d")
                                cell.number_format = 'yyyy-mm-dd'
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            except:
                                cell.number_format = '@'
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.value = ""
                            cell.number_format = '@'
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                # 为数值列设置格式
                for cell in ws['H'][1:]:  # H列是市场价值
                    if cell.value is not None and cell.value != "":
                        try:
                            float_val = float(cell.value)
                            cell.value = float_val
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        except (ValueError, TypeError):
                            cell.number_format = '@'
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # 为租金金额设置格式
                for cell in ws['R'][1:]:  # R列是租金金额
                    if cell.value is not None and cell.value != "":
                        try:
                            float_val = float(cell.value)
                            cell.value = float_val
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        except (ValueError, TypeError):
                            cell.number_format = '@'
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # 为托管合同金额设置格式
                for cell in ws['Y'][1:]:  # Y列是托管合同金额
                    if cell.value is not None and cell.value != "":
                        try:
                            float_val = float(cell.value)
                            cell.value = float_val
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        except (ValueError, TypeError):
                            cell.number_format = '@'
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # 为托管日期列设置条件格式
                trusteeship_date_columns = ['AA', 'AB', 'AC']  # 托管合同开始日期、结束日期、签署日期

                for col in trusteeship_date_columns:
                    for cell in ws[col][1:]:
                        if cell.value and str(cell.value).strip():
                            try:
                                datetime.strptime(str(cell.value).strip(), "%Y-%m-%d")
                                cell.number_format = 'yyyy-mm-dd'
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            except:
                                cell.number_format = '@'
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.value = ""
                            cell.number_format = '@'
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                # 设置所有数据行单元格为居中对齐
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if cell.alignment.horizontal is None:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # 设置边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 确定数据区域范围
                max_row = ws.max_row
                max_col = ws.max_column

                for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                    for cell in row:
                        cell.border = thin_border

                # 冻结标题行
                ws.freeze_panes = 'A2'

                # 保存文件
                wb.save(filepath)

                # 显示导出成功的详细信息
                messagebox.showinfo("导出成功",
                                    f"成功导出 {len(rows)} 条记录到:\n{filepath}\n\n"
                                    f"包含字段：{len(headers)} 个\n"
                                    f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

            except PermissionError:
                messagebox.showerror("错误", "无文件写入权限，请选择其他位置")
            except Exception as e:
                messagebox.showerror("错误", f"创建Excel文件失败: {str(e)}")
                import traceback
                traceback.print_exc()

        except Exception as e:
            messagebox.showerror("错误", f"导出过程中出错: {str(e)}")
            import traceback
            traceback.print_exc()

    def import_from_excel(self):
        """从Excel文件导入数据"""
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")],
            title="选择要导入的Excel文件"
        )

        if not filepath:
            return

        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            # 获取标题行
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

            # 验证Excel文件格式
            expected_headers = [
                "ID", "名称", "管理方式", "资产分类", "编号", "数量",
                "购置日期", "市场价值", "责任人", "地址", "状态",
                "租赁开始日期", "租赁结束日期", "承租方", "承租方联系方式",
                "承租方性质", "承租方用途", "租金金额", "租金交付方式", "公开招拍租情况",
                "产证情况", "产权单位", "建筑面积（平方米）",
                "托管合同类型", "托管合同金额", "托管合同相对方",
                "托管合同编号", "托管合同开始日期", "托管合同结束日期", "托管签署日期",
                "托管是否归档",
                "录入人", "所属部门", "备注"
            ]

            if headers != expected_headers:
                messagebox.showerror("错误",
                                     f"Excel文件格式不正确\n请使用本系统导出的Excel文件\n\n当前格式：{headers}\n预期格式：{expected_headers}")
                return

            # 读取数据
            imported_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                # 确保数据完整性
                row_data = list(row)
                imported_data.append(row_data)

            if not imported_data:
                messagebox.showwarning("警告", "Excel文件中没有有效数据")
                return

            conn = self.db.get_connection()
            c = conn.cursor()

            # 首先获取所有用户信息，创建用户名/姓名映射
            c.execute("SELECT id, username, full_name FROM users")
            users = c.fetchall()

            # 创建用户名和姓名的映射
            user_mapping = {}
            username_to_fullname = {}
            fullname_to_username = {}

            for user_id, username, full_name in users:
                user_mapping[username] = user_id
                user_mapping[full_name] = user_id
                username_to_fullname[username] = full_name
                fullname_to_username[full_name] = username

                # 特殊映射：系统管理员
                if username == "admin":
                    user_mapping["系统管理员"] = user_id
                    username_to_fullname["系统管理员"] = full_name if full_name else "系统管理员"
                    fullname_to_username["系统管理员"] = "admin"
                if full_name == "系统管理员":
                    user_mapping["系统管理员"] = user_id
                    username_to_fullname["系统管理员"] = full_name
                    fullname_to_username["系统管理员"] = "admin"

            # 导入数据
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            current_user = self.auth.get_current_user()
            current_username = current_user["username"] if current_user else "admin"

            imported_count = 0
            duplicate_count = 0
            errors = []

            # 日期清洗函数
            def clean_date(date_value):
                """清洗日期格式"""
                if not date_value or date_value == "":
                    return None

                date_str = str(date_value).strip()

                # 如果包含空格，取空格前的部分（YYYY-MM-DD HH:MM:SS）
                if ' ' in date_str:
                    date_str = date_str.split(' ')[0]

                # 尝试解析常见的日期格式
                try:
                    # 如果是 datetime 对象
                    if hasattr(date_value, 'strftime'):
                        return date_value.strftime("%Y-%m-%d")

                    # 尝试解析不同格式
                    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"]:
                        try:
                            dt = datetime.strptime(date_str, fmt)
                            return dt.strftime("%Y-%m-%d")
                        except ValueError:
                            continue

                    # 如果都不匹配，返回原始值
                    return date_str
                except Exception:
                    return None

            for i, row in enumerate(imported_data, start=2):  # 从第2行开始，i是行号
                try:
                    # 名称
                    asset_name = row[1] if len(row) > 1 else ""

                    # 管理方式
                    management_type = row[2] if len(row) > 2 else ""

                    # 验证管理方式是否有效
                    valid_management_types = ["自主管理", "租赁管理", "托管管理"]
                    if management_type not in valid_management_types:
                        errors.append(f"第{i}行: 管理方式 '{management_type}' 无效，应为：自主管理/租赁管理/托管管理")
                        continue

                    # 资产分类
                    category = row[3] if len(row) > 3 else ""

                    # 编号
                    asset_number = row[4] if len(row) > 4 else ""

                    # 数量
                    quantity_str = row[5] if len(row) > 5 else ""
                    try:
                        if not quantity_str or quantity_str == "":
                            quantity = 1
                        else:
                            quantity = int(float(quantity_str))  # 处理可能的浮点数
                    except ValueError:
                        quantity = 1
                        errors.append(f"第{i}行: 数量格式错误，已设为1")

                    # 购置日期
                    purchase_date_raw = row[6] if len(row) > 6 else None
                    purchase_date_db = clean_date(purchase_date_raw)

                    # 市场价值
                    market_value_raw = row[7] if len(row) > 7 else None
                    market_value_db = None
                    if market_value_raw and market_value_raw != "" and market_value_raw != 0:
                        try:
                            market_value_db = float(str(market_value_raw).replace('¥', '').replace(',', '').strip())
                        except ValueError:
                            market_value_db = None
                            errors.append(f"第{i}行: 市场价值格式错误，已设为空")

                    # 责任人
                    responsible_person = row[8] if len(row) > 8 else None

                    # 地址
                    location = row[9] if len(row) > 9 else ""
                    if not location or str(location).strip() == "":
                        location = "未填写地址"

                    # 状态
                    status = row[10] if len(row) > 10 else "正常使用"

                    # 租赁开始日期
                    lease_start_date_raw = row[11] if len(row) > 11 else None
                    lease_start_date_db = clean_date(lease_start_date_raw)

                    # 租赁结束日期
                    lease_end_date_raw = row[12] if len(row) > 12 else None
                    lease_end_date_db = clean_date(lease_end_date_raw)

                    # 租赁提醒天数
                    lease_reminder_days = 30

                    # 承租方名称
                    tenant_name = row[13] if len(row) > 13 else None

                    # 承租方联系方式
                    tenant_contact = row[14] if len(row) > 14 else None

                    # 承租方性质
                    tenant_nature = row[15] if len(row) > 15 else None

                    # 承租方用途
                    tenant_purpose = row[16] if len(row) > 16 else None

                    # 租金金额
                    rent_amount_raw = row[17] if len(row) > 17 else None
                    rent_amount_db = None
                    if rent_amount_raw and rent_amount_raw != "" and rent_amount_raw != 0:
                        try:
                            rent_amount_db = float(str(rent_amount_raw).replace('¥', '').replace(',', '').strip())
                        except ValueError:
                            rent_amount_db = None
                            errors.append(f"第{i}行: 租金金额格式错误，已设为空")

                    # 租金交付方式
                    rent_payment_method = row[18] if len(row) > 18 else None

                    # 公开招拍租情况
                    bidding_situation = row[19] if len(row) > 19 else None

                    # 产证情况
                    certificate_status = row[20] if len(row) > 20 else None

                    # 产权单位
                    property_unit = row[21] if len(row) > 21 else None

                    # 建筑面积
                    building_area = row[22] if len(row) > 22 else None

                    # 托管合同类型
                    trusteeship_contract_type = row[23] if len(row) > 23 else None

                    # 托管合同金额
                    trusteeship_contract_amount_raw = row[24] if len(row) > 24 else None
                    trusteeship_contract_amount_db = None
                    if trusteeship_contract_amount_raw and trusteeship_contract_amount_raw != "":
                        # 允许文本，不强制转换为数字
                        trusteeship_contract_amount_db = str(trusteeship_contract_amount_raw).strip()

                    # 托管合同相对方
                    trusteeship_counterparty = row[25] if len(row) > 25 else None

                    # 托管合同编号
                    trusteeship_contract_number = row[26] if len(row) > 26 else None

                    # 托管合同开始日期
                    trusteeship_start_date_raw = row[27] if len(row) > 27 else None
                    trusteeship_start_date_db = clean_date(trusteeship_start_date_raw)

                    # 托管合同结束日期
                    trusteeship_end_date_raw = row[28] if len(row) > 28 else None
                    trusteeship_end_date_db = clean_date(trusteeship_end_date_raw)

                    # 托管签署日期
                    trusteeship_sign_date_raw = row[29] if len(row) > 29 else None
                    trusteeship_sign_date_db = clean_date(trusteeship_sign_date_raw)

                    # 托管是否归档
                    trusteeship_is_archived = row[30] if len(row) > 30 else "否"

                    # 录入人 - 关键修改部分
                    created_by_excel = row[31] if len(row) > 31 else ""
                    created_by_db = current_username  # 默认使用当前用户

                    if created_by_excel and created_by_excel.strip():
                        # 先尝试将Excel中的录入人转换为数据库用户名
                        excel_input = created_by_excel.strip()

                        # 情况1: Excel中是用户名
                        if excel_input in username_to_fullname:
                            created_by_db = excel_input  # 直接使用用户名
                        # 情况2: Excel中是姓名
                        elif excel_input in fullname_to_username:
                            created_by_db = fullname_to_username[excel_input]
                        # 情况3: 特殊处理"系统管理员"
                        elif excel_input == "系统管理员":
                            if "admin" in username_to_fullname:
                                created_by_db = "admin"
                            else:
                                created_by_db = current_username
                                errors.append(
                                    f"第{i}行: 录入人 '系统管理员' 未找到对应账户，已使用 '{current_username}'")
                        # 情况4: 其他情况，使用当前用户
                        else:
                            errors.append(
                                f"第{i}行: 录入人 '{created_by_excel}' 未找到对应账户，已使用 '{current_username}'")
                    else:
                        # Excel中录入人为空，使用当前用户
                        created_by_db = current_username

                    # 所属部门
                    department_name = row[32] if len(row) > 32 else None

                    # 备注
                    notes = row[33] if len(row) > 33 else None

                    if not asset_name or not asset_number:
                        errors.append(f"第{i}行: 名称或编号不能为空")
                        continue

                    # 检查重复
                    c.execute("SELECT 1 FROM assets WHERE asset_number=?", (asset_number,))
                    if c.fetchone():
                        duplicate_count += 1
                        errors.append(f"第{i}行: 编号 '{asset_number}' 已存在，已跳过")
                        continue

                    # 获取部门ID
                    department_id = None
                    if department_name:
                        c.execute("SELECT id FROM departments WHERE name=?", (department_name,))
                        dept_result = c.fetchone()
                        if dept_result:
                            department_id = dept_result[0]
                        else:
                            # 查找部门名称映射
                            dept_name_mapping = {
                                "资产管理科": "资产管理科",
                                "资产管理科": "资产管理科",
                            }
                            mapped_name = dept_name_mapping.get(department_name, department_name)
                            c.execute("SELECT id FROM departments WHERE name=?", (mapped_name,))
                            dept_result = c.fetchone()
                            if dept_result:
                                department_id = dept_result[0]
                            else:
                                errors.append(f"第{i}行: 部门 '{department_name}' 不存在，将不设置部门")

                    # 如果部门ID为空且当前用户是普通员工或科长，自动使用当前用户的部门
                    if department_id is None and current_user and current_user['role'] in ['section_chief', 'staff']:
                        current_user_dept_id = current_user.get('department_id')
                        if current_user_dept_id:
                            department_id = current_user_dept_id

                    # 权限检查：非管理员只能导入自己部门的资产
                    if current_user and current_user['role'] in ['section_chief', 'staff']:
                        if department_id != current_user.get('department_id'):
                            errors.append(f"第{i}行: 您只能导入自己部门的资产")
                            continue

                    # 根据管理方式验证必要字段
                    if management_type == "租赁管理":
                        if not lease_end_date_db:
                            errors.append(f"第{i}行: 租赁管理资产必须填写租赁结束日期")
                            continue

                    elif management_type == "托管管理":
                        # 托管日期改为非必填，只验证格式
                        if trusteeship_end_date_db:
                            try:
                                datetime.strptime(trusteeship_end_date_db, "%Y-%m-%d")
                            except ValueError:
                                errors.append(f"第{i}行: 托管合同结束日期格式不正确")
                                continue

                        if trusteeship_start_date_db:
                            try:
                                datetime.strptime(trusteeship_start_date_db, "%Y-%m-%d")
                            except ValueError:
                                errors.append(f"第{i}行: 托管合同开始日期格式不正确")
                                continue

                    # 按照数据库实际的字段顺序（40个字段）
                    params = [
                        asset_name,  # 1. name
                        category,  # 2. category
                        management_type,  # 3. management_type
                        asset_number,  # 4. asset_number
                        quantity,  # 5. quantity
                        None,  # 6. model (Excel中没有)
                        purchase_date_db,  # 7. purchase_date
                        market_value_db,  # 8. market_value
                        responsible_person,  # 9. responsible_person
                        location,  # 10. location
                        status,  # 11. status
                        lease_start_date_db,  # 12. lease_start_date
                        lease_end_date_db,  # 13. lease_end_date
                        lease_reminder_days,  # 14. lease_reminder_days
                        tenant_name,  # 15. tenant_name
                        tenant_contact,  # 16. tenant_contact
                        None,  # 17. image_path1
                        None,  # 18. image_path2
                        None,  # 19. image_path3
                        notes,  # 20. notes
                        created_by_db,  # 21. created_by (使用转换后的用户名)
                        current_time,  # 22. created_at
                        department_id,  # 23. department_id
                        certificate_status,  # 24. certificate_status
                        property_unit,  # 25. property_unit
                        building_area,  # 26. building_area
                        trusteeship_contract_type,  # 27. trusteeship_contract_type
                        trusteeship_contract_amount_db,  # 28. trusteeship_contract_amount
                        trusteeship_counterparty,  # 29. trusteeship_counterparty
                        trusteeship_contract_number,  # 30. trusteeship_contract_number
                        trusteeship_start_date_db,  # 31. trusteeship_start_date
                        trusteeship_end_date_db,  # 32. trusteeship_end_date
                        trusteeship_sign_date_db,  # 33. trusteeship_sign_date
                        trusteeship_is_archived,  # 34. trusteeship_is_archived
                        tenant_nature,  # 35. tenant_nature
                        tenant_purpose,  # 36. tenant_purpose
                        rent_amount_db,  # 37. rent_amount
                        rent_payment_method,  # 38. rent_payment_method
                        bidding_situation  # 39. bidding_situation
                    ]

                    # 检查参数数量
                    if len(params) != 39:
                        errors.append(f"第{i}行: 参数数量错误 {len(params)} != 39")
                        continue

                    # 执行INSERT - 39个字段
                    c.execute('''INSERT INTO assets 
                                (name, category, management_type, asset_number, quantity, model,
                                 purchase_date, market_value, responsible_person,
                                 location, status, lease_start_date, lease_end_date,
                                 lease_reminder_days, tenant_name, tenant_contact,
                                 image_path1, image_path2, image_path3, notes,  -- notes位置变了
                                 created_by, created_at, department_id,
                                 certificate_status, property_unit, building_area,
                                 trusteeship_contract_type,
                                 trusteeship_contract_amount, trusteeship_counterparty,
                                 trusteeship_contract_number, trusteeship_start_date,
                                 trusteeship_end_date, trusteeship_sign_date,
                                 trusteeship_is_archived,
                                 tenant_nature, tenant_purpose, rent_amount, rent_payment_method, bidding_situation)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                              params)

                    imported_count += 1

                except Exception as e:
                    errors.append(f"第{i}行: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    continue

            conn.commit()
            conn.close()

            # 显示导入结果
            message = f"导入完成！\n\n"
            message += f"✓ 成功导入: {imported_count} 条记录\n"
            message += f"✗ 重复跳过: {duplicate_count} 条记录\n"
            message += f"📋 总计处理: {len(imported_data)} 条记录"

            if errors:
                error_count = len(errors)
                if error_count > 20:
                    message += f"\n\n❌ 发现 {error_count} 个错误（显示前20个）:"
                    for error in errors[:20]:
                        message += f"\n{error}"
                    message += f"\n... 还有 {error_count - 20} 个错误未显示"
                else:
                    message += f"\n\n❌ 发现 {error_count} 个错误:"
                    for error in errors:
                        message += f"\n{error}"

            messagebox.showinfo("导入结果", message)

            # 刷新显示
            if self.is_searching and self.last_search_keyword:
                self.search_assets()
            else:
                self.load_assets()

        except Exception as e:
            messagebox.showerror("错误", f"导入失败: {str(e)}\n\n请确保Excel文件未被其他程序打开")
            import traceback
            traceback.print_exc()

    def search_assets(self):
        """搜索资产"""
        keyword = self.search_entry.get().strip()
        if not keyword:
            messagebox.showwarning("提示", "请输入搜索关键词")
            return

        # 设置搜索状态
        self.is_searching = True
        self.last_search_keyword = keyword

        # 清空表格
        for item in self.tree.get_children():
            self.tree.delete(item)

        conn = self.db.get_connection()
        c = conn.cursor()

        # 构建查询语句
        query = '''SELECT a.id, a.name, a.management_type, a.category, a.asset_number, a.quantity,
                          a.purchase_date, a.market_value,
                          a.responsible_person, a.location, a.status,
                          a.lease_end_date, COALESCE(u.full_name, a.created_by) as display_name, 
                          a.department_id, d.name as department_name
                   FROM assets a
                   LEFT JOIN users u ON a.created_by = u.username
                   LEFT JOIN departments d ON a.department_id = d.id
                   WHERE (a.name LIKE ? OR a.category LIKE ? OR a.asset_number LIKE ? 
                         OR a.responsible_person LIKE ? OR a.location LIKE ? OR a.status LIKE ?
                         OR a.management_type LIKE ? OR d.name LIKE ? OR a.tenant_name LIKE ?
                         OR a.notes LIKE ?)'''

        search_param = f"%{keyword}%"
        params = [search_param] * 10  # 10个参数

        # 部门权限限制
        current_user = self.auth.get_current_user()
        if current_user['role'] in ['section_chief', 'staff']:
            query += " AND a.department_id = ?"
            params.append(current_user.get('department_id'))

        query += " ORDER BY a.id"

        c.execute(query, params)
        rows = c.fetchall()
        conn.close()

        if not rows:
            # 尝试搜索数字字段
            try:
                # 检查是否是数字
                if keyword.replace('.', '', 1).isdigit():
                    keyword_num = float(keyword)
                    conn = self.db.get_connection()
                    c = conn.cursor()

                    query = '''SELECT a.id, a.name, a.management_type, a.category, a.asset_number, a.quantity,
                                      a.purchase_date, a.market_value,
                                      a.responsible_person, a.location, a.status,
                                      a.lease_end_date, u.full_name as created_by,
                                      a.department_id, d.name as department_name
                               FROM assets a
                               LEFT JOIN users u ON a.created_by = u.username
                               LEFT JOIN departments d ON a.department_id = d.id
                               WHERE (a.market_value = ? OR a.quantity = ? OR a.id = ?)'''

                    params = [keyword_num, int(keyword_num) if '.' not in keyword else None,
                              int(keyword_num) if '.' not in keyword else None]

                    if current_user['role'] in ['section_chief', 'staff']:
                        query += " AND a.department_id = ?"
                        params.append(current_user.get('department_id'))

                    c.execute(query, params)
                    rows = c.fetchall()
                    conn.close()

                    if not rows:
                        self.status_var.set(f"未找到包含 '{keyword}' 的记录")
                        return
                else:
                    self.status_var.set(f"未找到包含 '{keyword}' 的记录")
                    return
            except:
                self.status_var.set(f"未找到包含 '{keyword}' 的记录")
                return

        # 显示搜索结果，使用显示序号
        display_id = 1
        for row in rows:
            # 创建一个新的列表，按照主界面要求的顺序
            formatted_row = [
                display_id,  # ID显示序号
                row[1],  # 名称
                row[2],  # 管理方式
                row[3],  # 资产分类
                row[4],  # 编号
                row[5],  # 数量
                row[6] if row[6] else "-", # 购置日期
                f"{row[7]:.2f}" if row[7] is not None and row[7] != 0 else "-",  # 市场价值
                row[8],  # 责任人
                row[9],  # 地址
                row[10],  # 状态
                self.get_lease_status(row[2], row[11]),  # 租赁状态
                row[12],  # 录入人
                row[14] if row[14] else "未分配部门"  # 部门名称
            ]

            self.tree.insert("", "end", values=formatted_row, tags=(f"asset_{row[0]}",))
            display_id += 1

        self.status_var.set(f"找到 {len(rows)} 条包含 '{keyword}' 的记录")

    def load_assets(self):
        """加载资产数据"""
        # 加载全部资产时清除搜索状态
        if not self.is_searching:
            self.last_search_keyword = ""

        try:
            for item in self.tree.get_children():
                self.tree.delete(item)

            conn = self.db.get_connection()
            c = conn.cursor()

            # 查询资产数据，包含部门信息
            query = '''SELECT a.id, a.name, a.management_type, a.category, a.asset_number, a.quantity,
                              a.purchase_date, a.market_value,
                              a.responsible_person, a.location, a.status,
                              a.lease_end_date,
                              COALESCE(u.full_name, a.created_by) as display_name,
                              a.department_id, d.name as department_name
                       FROM assets a
                       LEFT JOIN users u ON a.created_by = u.username
                       LEFT JOIN departments d ON a.department_id = d.id
                       '''

            conditions = []
            params = []

            # 部门权限控制
            current_user = self.auth.get_current_user()
            if current_user['role'] in ['section_chief', 'staff']:
                conditions.append("a.department_id = ?")
                params.append(current_user.get('department_id'))

            if conditions:
                query += " WHERE " + " AND ".join(conditions)

            query += " ORDER BY a.id"

            c.execute(query, params)
            rows = c.fetchall()

            # 使用显示序号而不是数据库ID
            display_id = 1
            for row in rows:
                formatted_row = [
                    display_id,  # 显示序号
                    row[1],  # 名称
                    row[2],  # 管理方式
                    row[3],  # 资产分类
                    row[4],  # 编号
                    row[5],  # 数量
                    row[6] if row[6] else "-",  # 购置日期
                    f"{row[7]:.2f}" if row[7] is not None and row[7] != 0 else "-",  # 市场价值
                    row[8],  # 责任人
                    row[9],  # 地址
                    row[10],  # 状态
                    self.get_lease_status(row[2], row[11]),  # 租赁状态
                    row[12],  # 录入人
                    row[14] if row[14] else "未分配部门"  # 部门名称
                ]

                self.tree.insert("", "end", values=formatted_row, tags=(f"asset_{row[0]}",))
                display_id += 1

            self.status_var.set(f"共加载 {len(rows)} 条资产记录")

        except Exception as e:
            messagebox.showerror("错误", f"加载资产失败: {e}")
            import traceback
            traceback.print_exc()
        finally:
            if 'conn' in locals():
                conn.close()

    def select_image(self, label_widget):
        """选择图片文件"""
        filepath = filedialog.askopenfilename(
            title="选择资产图片",
            filetypes=[
                ("图片文件", "*.png *.jpg *.jpeg *.gif *.bmp"),
                ("所有文件", "*.*")
            ]
        )

        if filepath:
            # 只显示文件名，不显示完整路径
            filename = os.path.basename(filepath)
            label_widget.config(text=filename)

            # 保存完整路径到自定义属性
            label_widget.image_full_path = filepath
            label_widget.is_image_selected = True  # 标记已选择图片
        else:
            # 用户取消选择，清空信息
            label_widget.config(text="未选择图片", foreground="gray")
            if hasattr(label_widget, 'image_full_path'):
                delattr(label_widget, 'image_full_path')
            if hasattr(label_widget, 'is_image_selected'):
                delattr(label_widget, 'is_image_selected')

    def save_asset(self, dialog, entries, asset_id=None):
        """保存资产信息"""

        # 自定义Entry取值方法
        def get_entry_value(entry):
            if entry is None or isinstance(entry, str):
                return ""
            elif hasattr(entry, 'get_value'):
                return entry.get_value()
            elif isinstance(entry, tk.Text):
                return entry.get("1.0", "end").strip()
            else:
                try:
                    return entry.get().strip()
                except AttributeError:
                    return ""

        # 坐标类型取值方法（从下拉框获取）
        def get_coord_type_value(entry):
            if entry is None:
                return "wgs84"
            try:
                val = entry.get()
                if "百度" in val:
                    return "bd09"
                elif "高德" in val or "火星" in val:
                    return "gcj02"
                else:
                    return "wgs84"
            except Exception:
                return "wgs84"

        # 构建数据字典
        data = {
            "name": entries["name"].get().strip(),
            "management_type": entries["management_type"].get().strip(),
            "category": entries["category"].get().strip(),
            "asset_number": entries["asset_number"].get().strip(),
            "quantity": entries["quantity"].get().strip(),
            "purchase_date": entries["purchase_date"].get().strip(),
            "market_value": entries["market_value"].get().strip(),
            "responsible_person": entries["responsible_person"].get().strip(),
            "location": entries["location"].get().strip(),
            "longitude": get_entry_value(entries.get("longitude")),
            "latitude": get_entry_value(entries.get("latitude")),
            "coord_type": get_coord_type_value(entries.get("coord_type")),
            "status": entries["status"].get(),
            "lease_start_date": get_entry_value(entries["lease_start_date"]),
            "lease_end_date": get_entry_value(entries["lease_end_date"]),
            "lease_reminder_days": get_entry_value(entries["lease_reminder_days"]),
            "tenant_name": get_entry_value(entries["tenant_name"]),
            "tenant_contact": get_entry_value(entries["tenant_contact"]),
            "tenant_nature": get_entry_value(entries["tenant_nature"]),
            "tenant_purpose": get_entry_value(entries["tenant_purpose"]),
            "rent_amount": get_entry_value(entries["rent_amount"]),
            "rent_payment_method": get_entry_value(entries["rent_payment_method"]),
            "bidding_situation": get_entry_value(entries["bidding_situation"]),
            "certificate_status": get_entry_value(entries["certificate_status"]),
            "property_unit": get_entry_value(entries["property_unit"]),
            "building_area": get_entry_value(entries["building_area"]),
            "notes": get_entry_value(entries["notes"]),
            "trusteeship_contract_type": get_entry_value(entries.get("trusteeship_contract_type")),
            "trusteeship_contract_amount": get_entry_value(entries.get("trusteeship_contract_amount")),
            "trusteeship_counterparty": get_entry_value(entries.get("trusteeship_counterparty", "")),
            "trusteeship_contract_number": get_entry_value(entries.get("trusteeship_contract_number", "")),
            "trusteeship_start_date": get_entry_value(entries.get("trusteeship_start_date", "")),
            "trusteeship_end_date": get_entry_value(entries.get("trusteeship_end_date", "")),
            "trusteeship_sign_date": get_entry_value(entries.get("trusteeship_sign_date", "")),
            "trusteeship_is_archived": entries.get("trusteeship_is_archived", "").get() if hasattr(
                entries.get("trusteeship_is_archived", ""), "get") else "",
        }



        # 处理部门信息
        dept_name = entries["department"].get().strip()
        if dept_name and dept_name != "未选择" and hasattr(self, 'department_ids'):
            data["department_id"] = self.department_ids.get(dept_name)
        else:
            data["department_id"] = None

        # 处理图片路径
        data["image_paths"] = []
        for i in range(3):
            if i < len(entries["image_paths"]):
                label_widget = entries["image_paths"][i]
                if hasattr(label_widget, 'image_full_path') and label_widget.image_full_path:
                    if os.path.exists(label_widget.image_full_path):
                        data["image_paths"].append(label_widget.image_full_path)
                    else:
                        data["image_paths"].append("")
                else:
                    data["image_paths"].append("")
            else:
                data["image_paths"].append("")

        # 必填校验
        required_fields = {
            "name": "资产名称",
            "management_type": "管理方式",
            "category": "资产分类",
            "asset_number": "资产编号",
            "responsible_person": "资产责任人",
        }

        for field, field_name in required_fields.items():
            field_value = data.get(field, "")
            print(f"  {field_name}({field}): '{field_value}', 类型: {type(field_value)}")

            # 更严格的验证
            if not field_value or str(field_value).strip() == "" or str(field_value).strip() in ["请输入资产名称",
                                                                                                 "如：GY2025"]:
                messagebox.showerror("错误", f"{field_name}不能为空")
                return


        # 租赁管理需要租赁信息
        if data["management_type"] == "租赁管理":
            if not data["lease_end_date"] or data["lease_end_date"] == "YYYY-MM-DD":
                messagebox.showerror("错误", "租赁管理资产必须填写租赁结束日期")
                return

        # 托管管理需要托管信息
        if data["management_type"] == "托管管理":
            # 如果有填写结束日期，验证格式
            if data["trusteeship_end_date"] and data["trusteeship_end_date"] not in ["YYYY-MM-DD", ""]:
                try:
                    datetime.strptime(data["trusteeship_end_date"], "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror("错误", "托管合同结束日期格式不正确，应为 YYYY-MM-DD")
                    return

            # 如果有填写开始日期，验证格式
            if data["trusteeship_start_date"] and data["trusteeship_start_date"] not in ["YYYY-MM-DD", ""]:
                try:
                    datetime.strptime(data["trusteeship_start_date"], "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror("错误", "托管合同开始日期格式不正确，应为 YYYY-MM-DD")
                    return

        # 类型与格式校验
        try:
            quantity_str = data["quantity"]
            if not quantity_str or quantity_str == "请输入整数":
                data["quantity"] = 1
            else:
                data["quantity"] = int(quantity_str)
        except ValueError:
            messagebox.showerror("错误", "数量格式不正确，请输入整数")
            return

        # 处理购置日期和市场价值
        purchase_date_value = data["purchase_date"]
        if not purchase_date_value or purchase_date_value == "YYYY-MM-DD":
            data["purchase_date"] = None
        else:
            data["purchase_date"] = purchase_date_value

        market_value_value = data["market_value"]
        if not market_value_value or market_value_value == "单位：元，可含小数":
            data["market_value"] = None
        else:
            try:
                market_value_str = market_value_value.replace('¥', '').replace(',', '').strip()
                data["market_value"] = float(market_value_str)
            except ValueError:
                messagebox.showerror("错误", "市场价值格式不正确，请输入数字")
                return

        # 处理租金金额
        rent_amount_value = data["rent_amount"]
        if not rent_amount_value or rent_amount_value == "单位：元，可含小数":
            data["rent_amount"] = None
        else:
            try:
                rent_amount_str = rent_amount_value.replace('¥', '').replace(',', '').strip()
                data["rent_amount"] = float(rent_amount_str)
            except ValueError:
                messagebox.showerror("错误", "租金金额格式不正确，请输入数字")
                return

        # 处理提醒天数
        try:
            reminder_str = data["lease_reminder_days"]
            if data["management_type"] == "租赁管理" and reminder_str and reminder_str != "1-365天":
                data["lease_reminder_days"] = int(reminder_str)
            elif not reminder_str or reminder_str == "1-365天":
                data["lease_reminder_days"] = 30
            else:
                data["lease_reminder_days"] = int(reminder_str) if reminder_str else 30
        except ValueError:
            messagebox.showerror("错误", "提醒天数格式不正确，请输入1-365之间的整数")
            return

        # 处理托管合同金额，允许文本输入
        trusteeship_contract_amount_value = data["trusteeship_contract_amount"]
        if not trusteeship_contract_amount_value or trusteeship_contract_amount_value == "单位：元":
            data["trusteeship_contract_amount"] = None
        else:
            # 允许文本输入，不强制转换为数字
            data["trusteeship_contract_amount"] = trusteeship_contract_amount_value.strip()

        # 处理地址字段
        if not data["location"] or data["location"] == "详细地址，如：XX市XX区XX路XX号":
            data["location"] = "未填写"
            print(f"地址为空，已设置为默认值: {data['location']}")

        # 权限验证
        current_user = self.auth.get_current_user()
        if current_user['role'] in ['section_chief', 'staff']:
            user_dept_id = current_user.get('department_id')
            if not asset_id:
                # 添加新资产，强制设置为用户部门
                data["department_id"] = user_dept_id
            else:
                # 编辑资产，检查部门权限
                conn_temp = self.db.get_connection()
                c_temp = conn_temp.cursor()
                c_temp.execute("SELECT department_id FROM assets WHERE id=?", (asset_id,))
                original_dept_id = c_temp.fetchone()
                conn_temp.close()

                if original_dept_id and original_dept_id[0] != user_dept_id:
                    messagebox.showerror("权限错误", "您只能编辑本部门的资产")
                    return

        # DB操作
        conn = self.db.get_connection()
        c = conn.cursor()

        try:
            # 确保资产编号唯一
            if asset_id:
                c.execute("SELECT 1 FROM assets WHERE asset_number=? AND id<>?", (data["asset_number"], asset_id))
            else:
                c.execute("SELECT 1 FROM assets WHERE asset_number=?", (data["asset_number"],))
            if c.fetchone():
                messagebox.showerror("错误", "资产编号已存在，请更换")
                return

            # 处理图片保存
            images_dir = os.path.join(project_root, "assets", "images")
            saved_image_paths = ["", "", ""]

            if asset_id:
                c.execute("SELECT image_path1, image_path2, image_path3 FROM assets WHERE id=?", (asset_id,))
                original_paths = c.fetchone()
                if original_paths:
                    saved_image_paths = list(original_paths)

            for i in range(3):
                current_path = data["image_paths"][i] if i < len(data["image_paths"]) else ""
                if not current_path or current_path == "未选择图片":
                    continue

                try:
                    os.makedirs(images_dir, exist_ok=True)
                    original_filename = os.path.basename(current_path)
                    safe_filename = "".join(
                        [c for c in original_filename if c.isalnum() or c in (' ', '.', '_')]).rstrip()

                    if asset_id:
                        filename = f"{asset_id}_{i}_{safe_filename}"
                    else:
                        filename = f"temp_{datetime.now().strftime('%Y%m%d%H%M%S')}_{i}_{safe_filename}"

                    dest_path = os.path.join(images_dir, filename)
                    shutil.copy2(current_path, dest_path)
                    saved_image_paths[i] = dest_path
                except Exception as e:
                    print(f"图片 {i + 1} 保存错误: {e}")

            # 处理房屋信息字段
            cert_status = None
            if data["certificate_status"] and data["certificate_status"].strip() and data[
                "certificate_status"] != "如房产证、土地证信息":
                cert_status = data["certificate_status"].strip()

            prop_unit = None
            if data["property_unit"] and data["property_unit"].strip() and data[
                "property_unit"] != "请输入产权单位名称":
                prop_unit = data["property_unit"].strip()

            building_area = None
            if data["building_area"] and data["building_area"].strip() and data["building_area"] != "单位：平方米":
                building_area = data["building_area"].strip()

            # 处理托管日期字段
            def clean_date(date_value):
                if not date_value or date_value == "":
                    return None
                date_str = str(date_value).strip()
                if ' ' in date_str:
                    date_str = date_str.split(' ')[0]
                return date_str

            trusteeship_start_date_db = clean_date(data["trusteeship_start_date"])
            trusteeship_end_date_db = clean_date(data["trusteeship_end_date"])
            trusteeship_sign_date_db = clean_date(data["trusteeship_sign_date"])

            print(f"🔄 开始数据库操作，asset_id: {asset_id}")

            if asset_id:  # 编辑模式
                print(f"📝 更新资产 ID: {asset_id}")

                update_params = [
                    data["name"], data["category"], data["management_type"], data["asset_number"], data["quantity"],
                    None,  # model字段
                    data["purchase_date"], data["market_value"], data["responsible_person"],
                    data["location"], data["status"],
                    data["lease_start_date"] if data["lease_start_date"] and data["lease_start_date"] not in [
                        "YYYY-MM-DD", "1-365天", "输入承租方全称", "电话或手机号"] else None,
                    data["lease_end_date"] if data["lease_end_date"] and data["lease_end_date"] not in [
                        "YYYY-MM-DD", "1-365天", "输入承租方全称", "电话或手机号"] else None,
                    data["lease_reminder_days"],
                    data["tenant_name"] if data["tenant_name"] and data["tenant_name"] not in [
                        "YYYY-MM-DD", "1-365天", "输入承租方全称", "电话或手机号"] else None,
                    data["tenant_contact"] if data["tenant_contact"] and data["tenant_contact"] not in [
                        "YYYY-MM-DD", "1-365天", "输入承租方全称", "电话或手机号"] else None,
                    data["tenant_nature"] if data["tenant_nature"] and data["tenant_nature"] != "-" else None,
                    data["tenant_purpose"] if data["tenant_purpose"] and data["tenant_purpose"] not in [
                        "YYYY-MM-DD", "1-365天", "输入承租方全称", "电话或手机号"] else None,
                    data["rent_payment_method"] if data["rent_payment_method"] and data[
                        "rent_payment_method"] not in ["YYYY-MM-DD", "1-365天", "输入承租方全称",
                                                       "电话或手机号"] else None,
                    data["bidding_situation"] if data["bidding_situation"] and data["bidding_situation"] not in [
                        "YYYY-MM-DD", "1-365天", "输入承租方全称", "电话或手机号"] else None,
                    # 房屋信息字段
                    cert_status, prop_unit, building_area,
                    data["rent_amount"],  # 租金金额字段
                    # 托管信息字段
                    data["trusteeship_contract_type"] if data["trusteeship_contract_type"] and data[
                        "trusteeship_contract_type"] not in ["请输入合同类型", ""] else None,
                    data["trusteeship_contract_amount"],
                    data["trusteeship_counterparty"] if data["trusteeship_counterparty"] and data[
                        "trusteeship_counterparty"] not in ["请输入合同相对方", ""] else None,
                    data["trusteeship_contract_number"] if data["trusteeship_contract_number"] and data[
                        "trusteeship_contract_number"] not in ["请输入合同编号", ""] else None,
                    trusteeship_start_date_db,
                    trusteeship_end_date_db,
                    trusteeship_sign_date_db,
                    data["trusteeship_is_archived"] if data["trusteeship_is_archived"] and data[
                        "trusteeship_is_archived"] != "" else "否",
                    # 图片路径
                    saved_image_paths[0], saved_image_paths[1], saved_image_paths[2],
                    # 备注
                    data["notes"],
                    # 部门ID
                    data["department_id"],
                    # 经纬度
                    float(data["longitude"]) if data.get("longitude") and data["longitude"] not in ["如：116.397428", ""] else None,
                    float(data["latitude"]) if data.get("latitude") and data["latitude"] not in ["如：39.90923", ""] else None,
                    # 坐标类型
                    get_coord_type_value(entries.get("coord_type")),
                    # WHERE条件
                    asset_id
                ]

                print(f"📋 更新参数数量: {len(update_params)}")

                c.execute(
                    '''UPDATE assets SET
                        name=?, category=?, management_type=?, asset_number=?, quantity=?,
                        model=?,
                        purchase_date=?, market_value=?, responsible_person=?,
                        location=?, status=?, lease_start_date=?, lease_end_date=?,
                        lease_reminder_days=?, tenant_name=?, tenant_contact=?,
                        tenant_nature=?, tenant_purpose=?, rent_payment_method=?, bidding_situation=?,
                        certificate_status=?, property_unit=?, building_area=?,
                        rent_amount=?,
                        trusteeship_contract_type=?,
                        trusteeship_contract_amount=?, trusteeship_counterparty=?,
                        trusteeship_contract_number=?, trusteeship_start_date=?,
                        trusteeship_end_date=?, trusteeship_sign_date=?,
                        trusteeship_is_archived=?,
                        image_path1=?, image_path2=?, image_path3=?, notes=?,
                        department_id=?, longitude=?, latitude=?, coord_type=?
                       WHERE id=?''',
                    update_params
                )

                print(f"✅ 更新资产成功")

            else:  # 新增模式
                print(f"➕ 新增资产")
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                created_by = self.auth.get_current_user()["username"]
                created_by = current_user["username"] if current_user and "username" in current_user else "admin"

                # 简化的插入参数，确保数量正确
                insert_params = [
                    data["name"], data["category"], data["management_type"], data["asset_number"], data["quantity"],
                    None,  # model字段
                    data["purchase_date"], data["market_value"], data["responsible_person"],
                    data["location"], data["status"],
                    data["lease_start_date"] if data["lease_start_date"] and data["lease_start_date"] not in [
                        "YYYY-MM-DD", "1-365天"] else None,
                    data["lease_end_date"] if data["lease_end_date"] and data["lease_end_date"] not in [
                        "YYYY-MM-DD", "1-365天"] else None,
                    data["lease_reminder_days"],
                    data.get("tenant_name"),
                    data.get("tenant_contact"),
                    data.get("tenant_nature"),
                    data.get("tenant_purpose"),
                    data.get("rent_amount"),
                    data.get("rent_payment_method"),
                    data.get("bidding_situation"),
                    # 房屋信息字段
                    cert_status, prop_unit, building_area,
                    # 托管信息字段
                    data["trusteeship_contract_type"] if data["trusteeship_contract_type"] and data[
                        "trusteeship_contract_type"] not in ["请输入合同类型", ""] else None,
                    data["trusteeship_contract_amount"],
                    data["trusteeship_counterparty"] if data["trusteeship_counterparty"] and data[
                        "trusteeship_counterparty"] not in ["请输入合同相对方", ""] else None,
                    data["trusteeship_contract_number"] if data["trusteeship_contract_number"] and data[
                        "trusteeship_contract_number"] not in ["请输入合同编号", ""] else None,
                    trusteeship_start_date_db,
                    trusteeship_end_date_db,
                    trusteeship_sign_date_db,
                    data["trusteeship_is_archived"] if data["trusteeship_is_archived"] and data[
                        "trusteeship_is_archived"] != "" else "否",
                    # 图片
                    saved_image_paths[0], saved_image_paths[1], saved_image_paths[2],
                    # 备注
                    data.get("notes", ""),
                    # 创建信息
                    created_by, current_time,
                    data.get("department_id"),
                    # 经纬度 + 坐标类型
                    float(data["longitude"]) if data.get("longitude") and data["longitude"] not in ["如：116.397428", ""] else None,
                    float(data["latitude"]) if data.get("latitude") and data["latitude"] not in ["如：39.90923", ""] else None,
                    data.get("coord_type", "wgs84")
                ]

                print(f"📋 插入参数数量: {len(insert_params)}")

                # 42个字段的INSERT语句
                c.execute(
                    '''INSERT INTO assets
                        (name, category, management_type, asset_number, quantity, model,
                         purchase_date, market_value, responsible_person,
                         location, status, lease_start_date, lease_end_date,
                         lease_reminder_days, tenant_name, tenant_contact,
                         tenant_nature, tenant_purpose, rent_amount, rent_payment_method, bidding_situation,
                         certificate_status, property_unit, building_area,
                         trusteeship_contract_type,
                         trusteeship_contract_amount, trusteeship_counterparty,
                         trusteeship_contract_number, trusteeship_start_date,
                         trusteeship_end_date, trusteeship_sign_date,
                         trusteeship_is_archived,
                         image_path1, image_path2, image_path3, notes,
                         created_by, created_at, department_id, longitude, latitude, coord_type)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    insert_params
                )

                new_id = c.lastrowid
                print(f"✅ 新增资产成功，ID: {new_id}")

            conn.commit()
            print(f"💾 数据库提交成功")

            # 验证保存结果
            c.execute("SELECT COUNT(*), MAX(id) FROM assets")
            result = c.fetchone()
            print(f"📊 验证: 总资产数={result[0]}, 最新ID={result[1]}")

            conn.close()

            # 关闭对话框
            dialog.destroy()

            # 强制刷新显示
            self.is_searching = False
            if hasattr(self, 'search_entry'):
                self.search_entry.delete(0, tk.END)

            self.load_assets()

            messagebox.showinfo("成功", "资产信息已保存")

        except Exception as e:
            conn.rollback()
            print(f"❌ 保存失败: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("错误", f"保存失败: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()

    def show_asset_detail(self, event):
        """显示资产详情 - 单页面展示所有信息"""
        selected_item = self.tree.selection()
        if not selected_item:
            return

        # 从标签中获取真实的数据库ID
        tags = self.tree.item(selected_item)["tags"]
        if not tags:
            return

        asset_id = int(tags[0].split("_")[1])

        conn = self.db.get_connection()
        c = conn.cursor()

        # 查看assets表的字段信息
        c.execute("PRAGMA table_info(assets)")
        columns = c.fetchall()
        print("数据库assets表字段信息:")
        for col in columns:
            print(f"  字段 {col[0]}: {col[1]} ({col[2]})")
        print(f"总字段数: {len(columns)}")

        # 查询语句
        c.execute('''
            SELECT 
                a.id, a.name, a.category, a.management_type, a.asset_number, a.quantity,
                a.model, a.purchase_date, a.market_value, a.responsible_person,
                a.location, a.status, a.lease_start_date, a.lease_end_date,
                a.lease_reminder_days, a.tenant_name, a.tenant_contact,
                a.tenant_nature, a.tenant_purpose, a.rent_payment_method, a.bidding_situation,
                a.certificate_status, a.property_unit, a.building_area,
                a.rent_amount,
                a.trusteeship_contract_type,
                a.trusteeship_contract_amount, a.trusteeship_counterparty,
                a.trusteeship_contract_number, a.trusteeship_start_date,
                a.trusteeship_end_date, a.trusteeship_sign_date,
                a.trusteeship_is_archived, 
                a.image_path1, a.image_path2, a.image_path3, a.notes,
                a.created_by,
                a.created_at, a.department_id, d.name as department_name
            FROM assets a
            LEFT JOIN departments d ON a.department_id = d.id
            WHERE a.id=?
        ''', (asset_id,))
        asset_data = c.fetchone()
        conn.close()

        if not asset_data:
            return

        # 创建对话框
        detail_dialog = tk.Toplevel(self.parent)
        detail_dialog.title(f"资产详情 - ID: {asset_data[0]}")
        detail_dialog.minsize(700, 550)
        detail_dialog.resizable(True, True)

        # 居中显示
        self.utils.center_window(detail_dialog)

        # 创建主框架和滚动条
        main_frame = ttk.Frame(detail_dialog)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 添加滚动条
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 明确的字段索引映射
        field_indices = {
            'id': 0,
            'name': 1,
            'category': 2,
            'management_type': 3,
            'asset_number': 4,
            'quantity': 5,
            'model': 6,
            'purchase_date': 7,
            'market_value': 8,
            'responsible_person': 9,
            'location': 10,
            'status': 11,
            'lease_start_date': 12,
            'lease_end_date': 13,
            'lease_reminder_days': 14,
            'tenant_name': 15,
            'tenant_contact': 16,
            'tenant_nature': 17,
            'tenant_purpose': 18,
            'rent_payment_method': 19,
            'bidding_situation': 20,
            'certificate_status': 21,
            'property_unit': 22,
            'building_area': 23,
            'rent_amount': 24,
            'trusteeship_contract_type': 25,
            'trusteeship_contract_amount': 26,
            'trusteeship_counterparty': 27,
            'trusteeship_contract_number': 28,
            'trusteeship_start_date': 29,
            'trusteeship_end_date': 30,
            'trusteeship_sign_date': 31,
            'trusteeship_is_archived': 32,
            'image_path1': 33,
            'image_path2': 34,
            'image_path3': 35,
            'notes': 36,
            'created_by': 37,
            'created_at': 38,
            'department_id': 39,
            'department_name': 40
        }

        # 获取管理方式
        management_type = asset_data[field_indices['management_type']]

        # 处理市场价值显示
        market_value = asset_data[field_indices['market_value']]
        if market_value is None or market_value == "" or market_value == 0:
            market_value_display = "-"
            market_value_numeric = 0
        else:
            try:
                market_value_numeric = float(market_value)
                if market_value_numeric >= 10000:
                    market_value_display = f"¥{market_value_numeric / 10000:,.2f} 万元"
                else:
                    market_value_display = f"¥{market_value_numeric:,.2f} 元"
            except (ValueError, TypeError):
                market_value_display = str(market_value)
                market_value_numeric = 0

        # 处理购置日期显示
        purchase_date = asset_data[field_indices['purchase_date']]
        purchase_date_display = purchase_date if purchase_date else "-"

        # 获取部门名称
        dept_name = asset_data[field_indices['department_name']]
        dept_name = dept_name if dept_name is not None else "未分配部门"

        # 基本信息区域
        basic_frame = tk.LabelFrame(scrollable_frame, text="基本信息", font=('微软雅黑', 12, 'bold'),
                                    padx=15, pady=10)
        basic_frame.pack(fill="x", pady=(0, 10))

        basic_fields = [
            ("资产名称", asset_data[field_indices['name']]),
            ("资产分类", asset_data[field_indices['category']]),
            ("管理方式", asset_data[field_indices['management_type']]),
            ("资产编号", asset_data[field_indices['asset_number']]),
            ("资产数量", str(asset_data[field_indices['quantity']])),
            ("购置日期", purchase_date_display),
            ("市场价值", market_value_display),
            ("资产责任人", asset_data[field_indices['responsible_person']]),
            ("资产地址", asset_data[field_indices['location']]),
            ("资产状态", asset_data[field_indices['status']] or "未设置"),
            ("所属部门", dept_name),
            ("录入人", asset_data[field_indices['created_by']] or "未知"),
            ("录入时间", asset_data[field_indices['created_at']] or "未知")
        ]

        for i, (field, value) in enumerate(basic_fields):
            row = i // 2
            col = (i % 2) * 2

            ttk.Label(basic_frame, text=f"{field}：",
                      font=('微软雅黑', 10, 'bold')).grid(
                row=row, column=col, padx=5, pady=5, sticky="e")

            if field == "资产地址":
                address_label = tk.Label(basic_frame, text=value,
                                         font=('微软雅黑', 10), fg="blue",
                                         cursor="hand2")
                address_label.grid(row=row, column=col + 1, padx=5, pady=5, sticky="w")
                address_label.bind("<Button-1>", lambda e, loc=value: self.open_map(loc))
            elif field == "市场价值":
                value_label = tk.Label(basic_frame, text=value,
                                       font=('微软雅黑', 10, 'bold'),
                                       foreground="red" if market_value_numeric >= 10000 else "black")
                value_label.grid(row=row, column=col + 1, padx=5, pady=5, sticky="w")
            elif field in ["资产责任人", "所属部门"]:
                # 粗体显示
                value_label = tk.Label(basic_frame, text=value,
                                       font=('微软雅黑', 10, 'bold'),
                                       foreground="black")
                value_label.grid(row=row, column=col + 1, padx=5, pady=5, sticky="w")
            else:
                value_label = ttk.Label(basic_frame, text=value, font=('微软雅黑', 10))
                value_label.grid(row=row, column=col + 1, padx=5, pady=5, sticky="w")

        # 房屋信息区域
        if asset_data[field_indices['category']] == "房屋资产":
            house_frame = tk.LabelFrame(scrollable_frame, text="房屋信息", font=('微软雅黑', 12, 'bold'),
                                        padx=15, pady=10)
            house_frame.pack(fill="x", pady=(0, 10))

            # 统一列配置
            house_frame.columnconfigure(0, minsize=100)  # 标签列统一宽度
            house_frame.columnconfigure(1, weight=1)  # 值列可扩展

            certificate_status = asset_data[field_indices['certificate_status']]
            property_unit = asset_data[field_indices['property_unit']]
            building_area = asset_data[field_indices['building_area']]

            certificate_status_display = "-" if not certificate_status else certificate_status
            property_unit_display = "-" if not property_unit else property_unit
            building_area_display = "-" if not building_area else building_area

            house_fields = [
                ("产证情况", certificate_status_display),
                ("产权单位", property_unit_display),
                ("建筑面积", f"{building_area_display} 平方米" if building_area_display != "-" else "-")
            ]

            for i, (field, value) in enumerate(house_fields):
                # 标签 - 右对齐，统一宽度
                ttk.Label(house_frame, text=f"{field}：",
                          font=('微软雅黑', 10, 'bold'),
                          width=10, anchor="e").grid(
                    row=i, column=0, padx=5, pady=8, sticky="e")

                if field == "产证情况":
                    # 多行文本框
                    text_frame = ttk.Frame(house_frame)
                    text_frame.grid(row=i, column=1, padx=5, pady=8, sticky="w")

                    notes_text = tk.Text(text_frame, width=40, height=3,
                                         font=('微软雅黑', 10), wrap="word")
                    if value == "-":
                        display_text = "-"
                    else:
                        display_text = value.replace(" 平方米", "")
                    notes_text.insert("1.0", display_text)
                    notes_text.config(state="disabled")

                    notes_scrollbar = ttk.Scrollbar(text_frame, orient="vertical",
                                                    command=notes_text.yview)
                    notes_text.configure(yscrollcommand=notes_scrollbar.set)

                    notes_text.pack(side="left", fill="both", expand=True)
                    notes_scrollbar.pack(side="right", fill="y")
                else:
                    value_label = ttk.Label(house_frame, text=value,
                                            font=('微软雅黑', 10), anchor="w")
                    value_label.grid(row=i, column=1, padx=5, pady=8, sticky="w")

        # 租赁信息区域
        if management_type == "租赁管理":
            lease_frame = tk.LabelFrame(scrollable_frame, text="租赁信息", font=('微软雅黑', 12, 'bold'),
                                        padx=15, pady=10)
            lease_frame.pack(fill="x", pady=(0, 10))

            # 统一列配置
            lease_frame.columnconfigure(0, minsize=100)  # 第一组标签列
            lease_frame.columnconfigure(2, minsize=100)  # 第二组标签列
            lease_frame.columnconfigure(1, weight=1)  # 第一组值列
            lease_frame.columnconfigure(3, weight=1)  # 第二组值列

            def format_rent_amount(amount):
                """格式化租金金额显示"""
                if amount is None or amount == "" or amount == 0:
                    return "-"
                try:
                    amount_float = float(amount)
                    if amount_float >= 10000:
                        return f"¥{amount_float / 10000:,.2f} 万元"
                    else:
                        return f"¥{amount_float:,.2f} 元"
                except (ValueError, TypeError):
                    return str(amount)

            lease_fields = [
                ("租赁开始日期", asset_data[field_indices['lease_start_date']] or "未设置"),
                ("租赁结束日期", asset_data[field_indices['lease_end_date']] or "未设置"),
                ("提前提醒天数", str(asset_data[field_indices['lease_reminder_days']] or 30) + "天"),
                ("承租方名称", asset_data[field_indices['tenant_name']] or "未设置"),
                ("承租方联系方式", asset_data[field_indices['tenant_contact']] or "未设置"),
                ("承租方性质", asset_data[field_indices['tenant_nature']] or "未设置"),
                ("承租方用途", asset_data[field_indices['tenant_purpose']] or "未设置"),
                ("租金金额", asset_data[field_indices['rent_amount']]),
                ("租金交付方式", asset_data[field_indices['rent_payment_method']] or "未设置"),
                ("公开招拍租情况", asset_data[field_indices['bidding_situation']] or "未设置")
            ]

            # 分两列显示
            for i, (field, value) in enumerate(lease_fields):
                row = i // 2
                col_offset = (i % 2) * 2  # 每两列为一组

                # 标签 - 右对齐，统一宽度
                ttk.Label(lease_frame, text=f"{field}：",
                          font=('微软雅黑', 10, 'bold'),
                          width=10, anchor="e").grid(
                    row=row, column=col_offset, padx=5, pady=8, sticky="e")

                if field == "租金金额":
                    # 格式化租金金额
                    rent_value = value
                    display_text = format_rent_amount(rent_value)
                    color = "black"
                    try:
                        if rent_value and rent_value != "未设置" and float(rent_value) != 0:
                            rent_float = float(rent_value)
                            if rent_float >= 10000:
                                color = "red"
                            elif rent_float >= 5000:
                                color = "orange"
                            elif rent_float > 0:
                                color = "green"
                    except (ValueError, TypeError):
                        color = "black"

                    value_label = tk.Label(lease_frame, text=display_text,
                                           font=('微软雅黑', 10, 'bold'),
                                           foreground=color, anchor="w")
                else:
                    value_label = ttk.Label(lease_frame, text=value,
                                            font=('微软雅黑', 10), anchor="w")

                value_label.grid(row=row, column=col_offset + 1, padx=5, pady=8, sticky="w")

            # 租赁状态显示
            lease_end_date = asset_data[field_indices['lease_end_date']]
            lease_status = self.get_lease_status("租赁管理", lease_end_date)
            status_row = (len(lease_fields) + 1) // 2

            ttk.Label(lease_frame, text="租赁状态：",
                      font=('微软雅黑', 10, 'bold'),
                      width=10, anchor="e").grid(
                row=status_row, column=0, padx=5, pady=8, sticky="e")

            status_color = "green"
            status_text = lease_status
            if lease_status == "已过期":
                status_color = "red"
            elif lease_status == "即将到期":
                status_color = "orange"

            status_label = tk.Label(lease_frame, text=status_text,
                                    font=('微软雅黑', 10, 'bold'), fg=status_color,
                                    anchor="w")
            status_label.grid(row=status_row, column=1, padx=5, pady=8, sticky="w")

            # 计算剩余天数
            if lease_end_date and lease_end_date not in ["未设置", "-"]:
                try:
                    end_date = datetime.strptime(lease_end_date, "%Y-%m-%d")
                    today = datetime.now()
                    days_left = (end_date - today).days

                    if days_left < 0:
                        days_text = f"已过期 {abs(days_left)} 天"
                        days_color = "red"
                    elif days_left <= 30:
                        days_text = f"剩余 {days_left} 天（即将到期）"
                        days_color = "orange"
                    else:
                        days_text = f"剩余 {days_left} 天"
                        days_color = "green"

                    ttk.Label(lease_frame, text="剩余天数：",
                              font=('微软雅黑', 10, 'bold'),
                              width=10, anchor="e").grid(
                        row=status_row, column=2, padx=5, pady=8, sticky="e")

                    days_label = tk.Label(lease_frame, text=days_text,
                                          font=('微软雅黑', 10, 'bold'), fg=days_color,
                                          anchor="w")
                    days_label.grid(row=status_row, column=3, padx=5, pady=8, sticky="w")
                except:
                    pass

        # 托管信息区域
        if management_type == "托管管理":
            trusteeship_frame = tk.LabelFrame(scrollable_frame, text="托管信息",
                                              font=('微软雅黑', 12, 'bold'),
                                              padx=15, pady=10)
            trusteeship_frame.pack(fill="x", pady=(0, 10))

            # 创建容器框架并统一列配置
            container = ttk.Frame(trusteeship_frame)
            container.pack(fill="x", expand=True)

            # 统一列配置
            container.columnconfigure(0, minsize=100)  # 第一组标签列
            container.columnconfigure(2, minsize=100)  # 第二组标签列
            container.columnconfigure(1, weight=1)  # 第一组值列
            container.columnconfigure(3, weight=1)  # 第二组值列

            # 托管信息字段
            trusteeship_fields = [
                ("合同类型", asset_data[field_indices['trusteeship_contract_type']]),
                ("合同金额", asset_data[field_indices['trusteeship_contract_amount']]),
                ("合同相对方", asset_data[field_indices['trusteeship_counterparty']]),
                ("合同编号", asset_data[field_indices['trusteeship_contract_number']]),
                ("合同开始日期", asset_data[field_indices['trusteeship_start_date']]),
                ("合同结束日期", asset_data[field_indices['trusteeship_end_date']]),
                ("签署日期", asset_data[field_indices['trusteeship_sign_date']]),
                ("是否归档", asset_data[field_indices['trusteeship_is_archived']]),
            ]

            # 使用grid布局，两列显示
            for i, (field, value) in enumerate(trusteeship_fields):
                row = i // 2  # 每行显示两个字段
                col_offset = (i % 2) * 2  # 标签在第0/2列，值在第1/3列

                # 标签 - 右对齐，统一宽度
                ttk.Label(container, text=f"{field}：",
                          font=('微软雅黑', 10, 'bold'),
                          width=10, anchor="e").grid(
                    row=row, column=col_offset, padx=5, pady=8, sticky="e")

                # 值显示
                if field == "合同金额":
                    # 格式化金额
                    if value is None or value == "" or value == 0:
                        display_text = "-"
                    else:
                        try:
                            amount_float = float(value)
                            if amount_float >= 10000:
                                display_text = f"¥{amount_float / 10000:,.2f} 万元"
                            else:
                                display_text = f"¥{amount_float:,.2f} 元"
                        except:
                            display_text = str(value)

                    # 设置颜色
                    color = "black"
                    try:
                        if value and float(value) > 0:
                            amount_float = float(value)
                            if amount_float >= 10000:
                                color = "red"
                            elif amount_float >= 5000:
                                color = "orange"
                            else:
                                color = "green"
                    except:
                        color = "black"

                    value_label = tk.Label(container, text=display_text,
                                           font=('微软雅黑', 10, 'bold'),
                                           fg=color, anchor="w")
                    value_label.grid(row=row, column=col_offset + 1, padx=5, pady=8, sticky="w")

                elif field == "合同相对方":
                    # 创建多行文本框
                    text_frame = ttk.Frame(container)
                    text_frame.grid(row=row, column=col_offset + 1, padx=5, pady=8, sticky="w")

                    text_widget = tk.Text(text_frame, width=30, height=3,
                                          font=('微软雅黑', 10), wrap="word")
                    text_widget.insert("1.0", value if value else "-")
                    text_widget.config(state="disabled")

                    scrollbar = ttk.Scrollbar(text_frame, orient="vertical",
                                              command=text_widget.yview)
                    text_widget.configure(yscrollcommand=scrollbar.set)

                    text_widget.pack(side="left", fill="both", expand=True)
                    scrollbar.pack(side="right", fill="y")

                elif field == "是否归档":
                    color = "green" if value == "是" else "red"
                    display_text = value if value else "否"
                    value_label = tk.Label(container, text=display_text,
                                           font=('微软雅黑', 10, 'bold'),
                                           fg=color, anchor="w")
                    value_label.grid(row=row, column=col_offset + 1, padx=5, pady=8, sticky="w")
                elif field in ["合同开始日期", "合同结束日期", "签署日期"]:
                    # 添加日期字段的特殊处理
                    display_text = value if value and value not in ["", "None", None] else "未填写"
                    value_label = ttk.Label(container, text=display_text,
                                            font=('微软雅黑', 10), anchor="w")
                    value_label.grid(row=row, column=col_offset + 1, padx=5, pady=8, sticky="w")
                else:
                    display_text = value if value else "-"
                    value_label = ttk.Label(container, text=display_text,
                                            font=('微软雅黑', 10), anchor="w")
                    value_label.grid(row=row, column=col_offset + 1, padx=5, pady=8, sticky="w")

            # 计算托管合同剩余天数
            trusteeship_end_date = asset_data[field_indices['trusteeship_end_date']]
            if trusteeship_end_date and trusteeship_end_date not in ["未设置", "-"]:
                try:
                    end_date = datetime.strptime(trusteeship_end_date, "%Y-%m-%d")
                    today = datetime.now()
                    days_left = (end_date - today).days

                    if days_left < 0:
                        status_text = f"合同已过期{abs(days_left)}天"
                        status_color = "red"
                    elif days_left <= 30:
                        status_text = f"合同即将到期，剩余{days_left}天"
                        status_color = "orange"
                    else:
                        status_text = f"合同正常，剩余{days_left}天"
                        status_color = "green"

                    # 添加合同状态
                    status_row = len(trusteeship_fields) // 2
                    if len(trusteeship_fields) % 2 != 0:
                        status_row += 1

                    ttk.Label(container, text="合同状态：",
                              font=('微软雅黑', 10, 'bold'),
                              width=10, anchor="e").grid(
                        row=status_row, column=0, padx=5, pady=(15, 5), sticky="e")

                    status_label = tk.Label(container, text=status_text,
                                            font=('微软雅黑', 10, 'bold'), fg=status_color)
                    status_label.grid(row=status_row, column=1, padx=5, pady=(15, 5), sticky="w")

                except Exception as e:
                    print(f"计算托管合同剩余天数出错: {e}")

        # 图片信息区域
        img_frame = tk.LabelFrame(scrollable_frame, text="资产图片", font=('微软雅黑', 12, 'bold'),
                                  padx=15, pady=10)
        img_frame.pack(fill="x", pady=(0, 10))

        # 获取图片路径
        image_paths = [
            asset_data[field_indices['image_path1']],
            asset_data[field_indices['image_path2']],
            asset_data[field_indices['image_path3']]
        ]

        # 创建图片容器
        img_container = ttk.Frame(img_frame)
        img_container.pack(fill="x")

        # 存储图片引用，防止被垃圾回收
        self.thumbnail_images = []

        has_images = False
        for i, img_path in enumerate(image_paths, 1):
            if img_path and os.path.exists(img_path):
                has_images = True
                try:
                    from PIL import Image, ImageTk

                    single_img_frame = ttk.LabelFrame(img_container, text=f"图片 {i}", padding=5)
                    single_img_frame.pack(side="left", padx=10, pady=5, fill="both", expand=True)

                    # 加载并创建缩略图
                    img = Image.open(img_path)
                    img.thumbnail((200, 200))  # 缩略图尺寸

                    photo = ImageTk.PhotoImage(img)
                    self.thumbnail_images.append(photo)

                    # 创建可点击的图片标签
                    img_label = tk.Label(single_img_frame, image=photo, cursor="hand2")
                    img_label.image = photo  # 保持引用
                    img_label.pack(pady=5)

                    # 存储完整图片路径用于放大
                    img_label.full_image_path = img_path
                    img_label.image_index = i

                    # 绑定点击事件 - 简化版本
                    img_label.bind("<Button-1>", lambda e, path=img_path: self.enlarge_image_simple(path))

                    filename_label = ttk.Label(single_img_frame, text=os.path.basename(img_path),
                                               font=('微软雅黑', 8))
                    filename_label.pack()

                except Exception as e:
                    error_frame = ttk.Frame(img_container)
                    error_frame.pack(side="left", padx=10, pady=5)

                    ttk.Label(error_frame, text=f"图片 {i} 加载失败",
                              font=('微软雅黑', 9), foreground="red").pack()
                    ttk.Label(error_frame, text=str(e), font=('微软雅黑', 8)).pack()

        if not has_images:
            ttk.Label(img_frame, text="暂无图片", font=('微软雅黑', 11),
                      foreground="gray").pack(pady=20)

        # 备注信息区域
        if asset_data[field_indices['notes']]:
            notes_frame = tk.LabelFrame(scrollable_frame, text="备注信息", font=('微软雅黑', 12, 'bold'),
                                        padx=15, pady=10)
            notes_frame.pack(fill="x", pady=(0, 10))

            notes_text = tk.Text(notes_frame, width=80, height=4,
                                 font=('微软雅黑', 10), wrap="word")
            notes_text.insert("1.0", asset_data[field_indices['notes']])
            notes_text.config(state="disabled")

            notes_scrollbar = ttk.Scrollbar(notes_frame, orient="vertical", command=notes_text.yview)
            notes_text.configure(yscrollcommand=notes_scrollbar.set)

            notes_text.pack(side="left", fill="both", expand=True)
            notes_scrollbar.pack(side="right", fill="y")

        # 按钮区域
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.pack(fill="x", pady=10)

        # 编辑按钮
        if self.auth.can_edit_asset(asset_data[field_indices['department_id']]):
            ttk.Button(button_frame, text="编辑资产",
                       command=lambda: self.edit_asset_from_detail(asset_id, detail_dialog)).pack(side="left", padx=5)

        ttk.Button(button_frame, text="关闭", command=detail_dialog.destroy).pack(side="right", padx=5)

        # 调整对话框尺寸
        def adjust_dialog_dimensions():
            try:
                scrollable_frame.update_idletasks()
                content_width = scrollable_frame.winfo_reqwidth()
                content_height = scrollable_frame.winfo_reqheight()

                screen_width = detail_dialog.winfo_screenwidth()
                screen_height = detail_dialog.winfo_screenheight()

                dialog_width = min(max(content_width + 40, 700), int(screen_width * 0.8))
                dialog_height = min(max(content_height + 60, 500), int(screen_height * 0.8))

                detail_dialog.geometry(f"{dialog_width}x{dialog_height}")
                self.utils.center_window(detail_dialog)

            except Exception as e:
                detail_dialog.geometry("900x600")
                self.utils.center_window(detail_dialog)

        detail_dialog.after(150, adjust_dialog_dimensions)

        # 添加鼠标滚轮支持
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")

        detail_dialog.bind("<Destroy>", _unbind_mousewheel)

    def enlarge_image_simple(self, image_path):
        """简化版图片放大查看"""
        if not image_path or not os.path.exists(image_path):
            messagebox.showerror("错误", "图片文件不存在或已被移动")
            return

        try:
            from PIL import Image, ImageTk

            # 创建放大图片对话框
            enlarge_dialog = tk.Toplevel(self.parent)
            enlarge_dialog.title(f"放大查看图片")
            enlarge_dialog.geometry("800x600")

            # 设置对话框属性
            enlarge_dialog.grab_set()
            enlarge_dialog.focus_set()

            # 居中显示
            self.utils.center_window(enlarge_dialog)

            # 创建主框架
            main_frame = ttk.Frame(enlarge_dialog, padding=10)
            main_frame.pack(fill="both", expand=True)

            # 创建Canvas用于显示大图
            canvas_frame = ttk.Frame(main_frame)
            canvas_frame.pack(fill="both", expand=True)

            canvas = tk.Canvas(canvas_frame, bg='white', highlightthickness=1, highlightbackground='gray')

            # 水平和垂直滚动条
            h_scrollbar = ttk.Scrollbar(canvas_frame, orient="horizontal", command=canvas.xview)
            v_scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)

            canvas.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

            # 网格布局
            canvas.grid(row=0, column=0, sticky="nsew")
            v_scrollbar.grid(row=0, column=1, sticky="ns")
            h_scrollbar.grid(row=1, column=0, sticky="ew")

            canvas_frame.grid_rowconfigure(0, weight=1)
            canvas_frame.grid_columnconfigure(0, weight=1)

            # 加载图片
            img = Image.open(image_path)
            original_width, original_height = img.size

            # 计算适合窗口的尺寸（保持纵横比）
            max_width = 750
            max_height = 500

            if original_width > max_width or original_height > max_height:
                ratio = min(max_width / original_width, max_height / original_height)
                display_width = int(original_width * ratio)
                display_height = int(original_height * ratio)
                img = img.resize((display_width, display_height), Image.LANCZOS)
            else:
                display_width = original_width
                display_height = original_height

            # 创建PhotoImage
            photo = ImageTk.PhotoImage(img)

            # 存储引用，防止被垃圾回收
            if not hasattr(self, 'enlarged_images'):
                self.enlarged_images = []
            self.enlarged_images.append(photo)

            # 在Canvas上显示图片
            canvas.create_image(0, 0, anchor="nw", image=photo)
            canvas.config(scrollregion=canvas.bbox("all"))

            # 底部信息栏
            bottom_frame = ttk.Frame(main_frame)
            bottom_frame.pack(fill="x", pady=(10, 0))

            # 原图尺寸信息
            size_info = f"原图尺寸: {original_width} × {original_height} 像素"
            if display_width != original_width or display_height != original_height:
                size_info += f" | 显示尺寸: {display_width} × {display_height} 像素"

            ttk.Label(bottom_frame, text=size_info, font=('微软雅黑', 9)).pack(side="left", padx=5)

            # 文件名
            filename = os.path.basename(image_path)
            ttk.Label(bottom_frame, text=f"文件名: {filename}", font=('微软雅黑', 9)).pack(side="left", padx=20)

            # 关闭按钮
            ttk.Button(bottom_frame, text="关闭", command=enlarge_dialog.destroy).pack(side="right")

            # 添加键盘快捷键
            def on_key_press(event):
                if event.keysym == 'Escape':
                    enlarge_dialog.destroy()

            enlarge_dialog.bind("<KeyPress>", on_key_press)

        except Exception as e:
            messagebox.showerror("错误", f"无法加载图片: {str(e)}")
            import traceback
            traceback.print_exc()

    def edit_asset(self):
        """编辑资产对话框"""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一条记录")
            return

        # 从标签中获取真实的数据库ID
        tags = self.tree.item(selected_item)["tags"]
        if not tags:
            messagebox.showerror("错误", "无法获取资产ID")
            return

        asset_id = int(tags[0].split("_")[1])

        # 获取选中资产的部门ID
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT department_id FROM assets WHERE id=?", (asset_id,))
        result = c.fetchone()
        conn.close()

        asset_dept_id = result[0] if result else None

        # 修改权限检查逻辑
        if not self.auth.can_edit_asset(asset_dept_id):
            # 获取资产部门名称用于提示
            conn = self.db.get_connection()
            c = conn.cursor()
            c.execute("SELECT d.name FROM assets a LEFT JOIN departments d ON a.department_id=d.id WHERE a.id=?",
                      (asset_id,))
            dept_result = c.fetchone()
            conn.close()

            asset_dept_name = dept_result[0] if dept_result else "未知部门"
            user_dept_name = self.get_department_name(self.auth.get_current_user().get('department_id'))

            messagebox.showerror("权限不足", f"您只能编辑本部门({user_dept_name})的资产\n此资产属于：{asset_dept_name}")
            return

        # 获取资产的完整数据用于填充表单
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM assets WHERE id=?", (asset_id,))
        asset_data = c.fetchone()
        conn.close()

        if asset_data:
            # 将数据库ID添加到数据列表的开头，用于对话框
            item_data = [asset_id] + list(asset_data[1:16])
            self.asset_dialog("编辑资产", item_data)

    def edit_asset_from_detail(self, asset_id, detail_dialog):
        """从详情对话框编辑资产"""
        # 关闭详情对话框
        detail_dialog.destroy()

        # 获取选中资产的部门ID用于权限检查
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT department_id FROM assets WHERE id=?", (asset_id,))
        result = c.fetchone()
        conn.close()

        asset_dept_id = result[0] if result else None

        # 权限检查逻辑
        if not self.auth.can_edit_asset(asset_dept_id):
            # 获取资产部门名称用于提示
            conn = self.db.get_connection()
            c = conn.cursor()
            c.execute("SELECT d.name FROM assets a LEFT JOIN departments d ON a.department_id=d.id WHERE a.id=?",
                      (asset_id,))
            dept_result = c.fetchone()
            conn.close()

            asset_dept_name = dept_result[0] if dept_result else "未知部门"
            user_dept_name = self.get_department_name(self.auth.get_current_user().get('department_id'))

            messagebox.showerror("权限不足", f"您只能编辑本部门({user_dept_name})的资产\n此资产属于：{asset_dept_name}")
            return

        # 清除当前选择
        for item in self.tree.selection():
            self.tree.selection_remove(item)

        # 查找并选择对应的行
        for item in self.tree.get_children():
            tags = self.tree.item(item)["tags"]
            if tags:
                item_asset_id = int(tags[0].split("_")[1])
                if item_asset_id == asset_id:
                    self.tree.selection_set(item)
                    break

        # 获取资产的完整数据用于填充表单
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM assets WHERE id=?", (asset_id,))
        asset_data = c.fetchone()
        conn.close()

        if asset_data:
            # 将数据库ID添加到数据列表的开头，用于对话框
            item_data = [asset_id] + list(asset_data[1:16])
            self.asset_dialog("编辑资产", item_data)


    def open_map(self, location):
        """打开地图选择对话框"""
        if not location or location == "未设置":
            messagebox.showinfo("提示", "地址信息为空")
            return

        # 创建地图选择对话框
        map_dialog = tk.Toplevel(self.parent)
        map_dialog.title("选择地图服务")
        map_dialog.geometry("400x200")
        map_dialog.resizable(False, False)
        map_dialog.grab_set()

        # 居中显示
        self.utils.center_window(map_dialog)

        # 主框架
        main_frame = ttk.Frame(map_dialog, padding=25)
        main_frame.pack(fill="both", expand=True)

        # 提示信息
        ttk.Label(main_frame, text="请选择地图服务查看位置：",
                  font=('微软雅黑', 12, 'bold')).pack(pady=(0, 10))

        # 地址显示
        address_frame = ttk.Frame(main_frame)
        address_frame.pack(fill="x", pady=(0, 20))

        ttk.Label(address_frame, text="地址：",
                  font=('微软雅黑', 10, 'bold')).pack(side="left")

        address_label = ttk.Label(address_frame, text=location,
                                  font=('微软雅黑', 10), foreground="blue",
                                  wraplength=300)
        address_label.pack(side="left", padx=(5, 0))

        # 按钮框架
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)

        def open_baidu_map():
            """打开百度地图并自动搜索地址"""
            try:
                import webbrowser
                import urllib.parse

                # 百度地图正确的URL格式
                encoded_location = urllib.parse.quote(location)

                # 使用百度地图Web版
                url = f"https://api.map.baidu.com/geocoder?address={encoded_location}&output=html"

                webbrowser.open(url)
                map_dialog.destroy()
            except Exception as e:
                messagebox.showerror("错误", f"无法打开百度地图：{str(e)}")

        def open_gaode_map():
            """打开高德地图并自动搜索地址"""
            try:
                import webbrowser
                import urllib.parse
                # 高德地图URL格式
                encoded_location = urllib.parse.quote(location)
                url = f"https://uri.amap.com/search?keyword={encoded_location}"
                webbrowser.open(url)
                map_dialog.destroy()
            except Exception as e:
                messagebox.showerror("错误", f"无法打开高德地图：{str(e)}")

        # 创建更大的按钮
        baidu_btn = tk.Button(btn_frame, text="百度地图",
                              command=open_baidu_map,
                              font=('微软雅黑', 11, 'bold'),
                              bg="#3385FF", fg="white",
                              width=12, height=2,
                              relief="raised", bd=2,
                              cursor="hand2")
        baidu_btn.pack(side="left", padx=15)

        def on_enter_baidu(e):
            baidu_btn.config(bg="#2971E6")

        def on_leave_baidu(e):
            baidu_btn.config(bg="#3385FF")

        baidu_btn.bind("<Enter>", on_enter_baidu)
        baidu_btn.bind("<Leave>", on_leave_baidu)

        gaode_btn = tk.Button(btn_frame, text="高德地图",
                              command=open_gaode_map,
                              font=('微软雅黑', 11, 'bold'),
                              bg="#0081FF", fg="white",
                              width=12, height=2,
                              relief="raised", bd=2,
                              cursor="hand2")
        gaode_btn.pack(side="left", padx=15)

        def on_enter_gaode(e):
            gaode_btn.config(bg="#0066CC")

        def on_leave_gaode(e):
            gaode_btn.config(bg="#0081FF")

        gaode_btn.bind("<Enter>", on_enter_gaode)
        gaode_btn.bind("<Leave>", on_leave_gaode)

        # 取消按钮
        cancel_frame = ttk.Frame(main_frame)
        cancel_frame.pack(pady=(10, 0))

        ttk.Button(cancel_frame, text="取消",
                   command=map_dialog.destroy,
                   width=10).pack()
