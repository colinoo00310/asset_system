import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import openpyxl
import sys
import os

# 添加项目根目录到路径（兼容开发和打包环境）
def get_project_root():
    """获取项目根目录"""
    if getattr(sys, 'frozen', False):
        if hasattr(sys, '_MEIPASS'):
            return sys._MEIPASS
        else:
            return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

project_root = get_project_root()
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from utils.utils import Utils


class DepartmentManagement:
    def __init__(self, parent, db, auth):
        self.parent = parent
        self.db = db
        self.auth = auth
        self.utils = Utils()

        self.clear_content()
        self.create_department_interface()
        self.load_departments()

    def clear_content(self):
        """清除内容区域"""
        for widget in self.parent.winfo_children():
            widget.destroy()

    def create_department_interface(self):
        """创建部门管理界面"""
        # 顶部工具栏
        toolbar = ttk.Frame(self.parent, padding=10)
        toolbar.pack(fill="x")

        # 获取当前用户
        current_user = self.auth.get_current_user()

        # 根据权限和角色显示按钮
        if current_user['role'] == 'admin':
            # 管理员显示所有按钮
            if self.auth.has_permission("departments", "add"):
                ttk.Button(toolbar, text="添加部门", command=self.add_department).pack(side="left", padx=5)
            if self.auth.has_permission("departments", "edit"):
                ttk.Button(toolbar, text="编辑部门", command=self.edit_department).pack(side="left", padx=5)
            if self.auth.has_permission("departments", "delete"):
                ttk.Button(toolbar, text="删除部门", command=self.delete_department).pack(side="left", padx=5)
            ttk.Button(toolbar, text="管理员工", command=self.manage_employees).pack(side="left", padx=5)

        elif current_user['role'] == 'section_chief':
            # 科长只显示"编辑部门"和"管理员工"按钮
            ttk.Button(toolbar, text="编辑部门", command=self.edit_department).pack(side="left", padx=5)
            ttk.Button(toolbar, text="管理员工", command=self.manage_employees).pack(side="left", padx=5)
        else:
            # staff角色只显示"查看员工"按钮
            ttk.Button(toolbar, text="查看员工", command=self.view_employees).pack(side="left", padx=5)

        # 导出、导入、刷新等按钮显示
        if self.auth.has_permission("departments", "export"):
            ttk.Button(toolbar, text="导出Excel", command=self.export_departments_to_excel).pack(side="left", padx=5)

        # 只有admin有导入权限
        if current_user['role'] == 'admin' and self.auth.has_permission("departments", "import"):
            ttk.Button(toolbar, text="导入Excel", command=self.import_departments_from_excel).pack(side="left", padx=5)

        ttk.Button(toolbar, text="刷新", command=self.load_departments).pack(side="left", padx=5)
        ttk.Button(toolbar, text="导出模板", command=self.export_template).pack(side="left", padx=5)

        # 搜索区域
        search_frame = ttk.Frame(self.parent, padding=10)
        search_frame.pack(fill="x")

        ttk.Label(search_frame, text="搜索:").pack(side="left")
        self.dept_search_entry = ttk.Entry(search_frame, width=40)
        self.dept_search_entry.pack(side="left", padx=5)
        ttk.Button(search_frame, text="搜索", command=self.search_departments).pack(side="left", padx=5)
        # 重置按钮，添加清空搜索框功能
        ttk.Button(search_frame, text="重置", command=self.reset_search).pack(side="left", padx=5)

        # 数据显示表格
        tree_frame = ttk.Frame(self.parent)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.dept_tree = ttk.Treeview(tree_frame,
                                      columns=("ID", "部门名称", "负责人", "联系方式", "地址"),
                                      show="headings")

        # 设置列宽和标题
        columns = {
            "ID": {"width": 40, "anchor": "center"},
            "部门名称": {"width": 120, "anchor": "center"},
            "负责人": {"width": 80, "anchor": "center"},
            "联系方式": {"width": 120, "anchor": "center"},
            "地址": {"width": 200, "anchor": "center"}
        }

        for col, settings in columns.items():
            self.dept_tree.column(col, **settings)
            self.dept_tree.heading(col, text=col)

        # 绑定双击事件
        self.dept_tree.bind("<Double-1>", self.show_department_detail)

        # 滚动条
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.dept_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.dept_tree.xview)
        self.dept_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.dept_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 状态栏
        status_bar = ttk.Frame(self.parent, height=25)
        status_bar.pack(fill="x", side="bottom")
        self.status_var = tk.StringVar()
        self.status_var.set(
            f"欢迎, {self.auth.get_current_user()['full_name'] or self.auth.get_current_user()['username']} | 角色: {self.auth.get_current_user()['role']}")
        ttk.Label(status_bar, textvariable=self.status_var, relief="sunken", padding=2).pack(fill="x")

    def reset_search(self):
        """重置搜索框并刷新部门列表"""
        self.dept_search_entry.delete(0, tk.END)
        self.load_departments()

    def view_employees(self):
        """查看员工对话框（只读模式）"""
        selected_item = self.dept_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一个部门")
            return

        dept_id = self.dept_tree.item(selected_item)["values"][0]
        dept_name = self.dept_tree.item(selected_item)["values"][1]

        dialog = tk.Toplevel(self.parent)
        dialog.title(f"查看员工 - {dept_name}")
        dialog.geometry("600x400")
        dialog.grab_set()

        # 居中显示
        self.utils.center_window(dialog)

        # 显示当前部门信息
        dept_info_frame = ttk.Frame(dialog, padding=5)
        dept_info_frame.pack(fill="x")

        ttk.Label(dept_info_frame, text=f"当前部门: {dept_name}",
                  font=('微软雅黑', 12, 'bold')).pack(anchor="w")

        # 员工表格（只读）
        tree_frame = ttk.Frame(dialog)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        emp_tree = ttk.Treeview(tree_frame, columns=("序号", "姓名", "职位", "联系方式", "排序号"), show="headings")

        # 设置列
        columns = {
            "序号": {"width": 50, "anchor": "center"},
            "姓名": {"width": 100, "anchor": "center"},
            "职位": {"width": 120, "anchor": "center"},
            "联系方式": {"width": 150, "anchor": "center"},
            "排序号": {"width": 80, "anchor": "center"}
        }

        for col, settings in columns.items():
            emp_tree.column(col, **settings)
            emp_tree.heading(col, text=col)

        # 滚动条
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=emp_tree.yview)
        emp_tree.configure(yscrollcommand=vsb.set)

        emp_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 加载员工数据
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT name, position, contact, sort_order FROM employees WHERE department_id=? ORDER BY sort_order, id",
                  (dept_id,))
        rows = c.fetchall()
        conn.close()

        # 显示排序序号
        for index, row in enumerate(rows, 1):
            display_values = (index, row[0], row[1], row[2], row[3] if row[3] else index)
            emp_tree.insert("", "end", values=display_values)

        # 关闭按钮
        ttk.Button(dialog, text="关闭", command=dialog.destroy).pack(pady=10)

    def load_departments(self):
        """加载部门数据"""
        try:
            for item in self.dept_tree.get_children():
                self.dept_tree.delete(item)

            conn = self.db.get_connection()
            c = conn.cursor()

            c.execute("SELECT id, name, manager, contact, address FROM departments ORDER BY id")
            rows = c.fetchall()
            conn.close()

            # 显示连续序号而不是数据库ID，但用标签存储真实ID
            for index, row in enumerate(rows, 1):
                # 顺序：序号, 部门名称, 负责人, 联系方式, 地址
                display_values = (index, row[1], row[2], row[3], row[4])
                self.dept_tree.insert("", "end", values=display_values, tags=(f"dept_{row[0]}",))

            if hasattr(self, 'status_var'):
                self.status_var.set(f"共加载 {len(rows)} 条部门记录")

        except Exception as e:
            messagebox.showerror("错误", f"加载部门失败: {e}")
        finally:
            if 'conn' in locals():
                conn.close()

    def search_departments(self):
        """搜索部门"""
        keyword = self.dept_search_entry.get().strip()
        if not keyword:
            messagebox.showwarning("提示", "请输入搜索关键词")
            return

        for item in self.dept_tree.get_children():
            self.dept_tree.delete(item)

        conn = self.db.get_connection()
        c = conn.cursor()

        # 修改查询字段顺序：id, name, manager, contact, address
        query = '''SELECT id, name, manager, contact, address 
                   FROM departments 
                   WHERE name LIKE ? OR address LIKE ? OR contact LIKE ? OR manager LIKE ?
                   ORDER BY id'''

        search_param = f"%{keyword}%"
        c.execute(query, (search_param, search_param, search_param, search_param))
        rows = c.fetchall()
        conn.close()

        # 使用显示序号而不是数据库ID
        display_id = 1
        for row in rows:
            display_values = (display_id, row[1], row[2], row[3], row[4])
            self.dept_tree.insert("", "end", values=display_values, tags=(f"dept_{row[0]}",))
            display_id += 1

        self.status_var.set(f"找到 {len(rows)} 条匹配记录")

    def add_department(self):
        """添加部门对话框"""
        self.department_dialog("添加部门")

    def edit_department(self):
        """编辑部门对话框"""
        selected_item = self.dept_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一条记录")
            return

        # 获取真实的数据库ID
        tags = self.dept_tree.item(selected_item, "tags")
        if not tags:
            messagebox.showerror("错误", "无法获取部门ID")
            return

        dept_id = int(tags[0].split("_")[1])

        # 获取当前用户信息
        current_user = self.auth.get_current_user()

        # 如果是科长，检查是否是自己所在的部门
        if current_user['role'] == 'section_chief':
            user_dept_id = current_user.get('department_id')

            # 如果用户没有关联部门，提示错误
            if not user_dept_id:
                messagebox.showerror("错误", "您的账号未分配部门，请联系管理员")
                return

            # 检查是否是自己所在的部门
            if user_dept_id != dept_id:
                # 获取用户部门名称
                conn = self.db.get_connection()
                c = conn.cursor()
                c.execute("SELECT name FROM departments WHERE id=?", (user_dept_id,))
                user_dept_name = c.fetchone()
                conn.close()

                user_dept_name_display = user_dept_name[0] if user_dept_name else "未知部门"
                messagebox.showerror("权限不足", f"您只能编辑自己所在的部门：{user_dept_name_display}")
                return

        # 从数据库获取真实数据，字段顺序为：id, name, manager, contact, address
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT id, name, manager, contact, address FROM departments WHERE id=?", (dept_id,))
        real_data = c.fetchone()
        conn.close()

        if not real_data:
            messagebox.showerror("错误", "未找到部门信息")
            return

        self.department_dialog("编辑部门", real_data)

    def department_dialog(self, title, data=None):
        """部门添加/编辑对话框"""
        dialog = tk.Toplevel(self.parent)
        dialog.title(title)
        dialog.grab_set()

        # 居中显示
        self.utils.center_window(dialog)

        # 主框架
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack()

        entries = {}

        # 获取当前用户
        current_user = self.auth.get_current_user()

        # 如果是编辑模式，先确保数据不为空
        if data:
            # data数据顺序：id, name, manager, contact, address
            if len(data) < 5:
                messagebox.showerror("错误", "部门数据不完整")
                dialog.destroy()
                return

            # 确保有部门名称
            if not data[1]:
                messagebox.showerror("错误", "部门名称不能为空")
                dialog.destroy()
                return

        # 表单字段，按顺序显示：部门名称、负责人、联系方式、地址
        fields = [
            ("部门名称", "name", 0),
            ("负责人", "manager", 1),
            ("联系方式", "contact", 2),
            ("地址", "address", 3)
        ]

        for label, field, row in fields:
            ttk.Label(main_frame, text=label + ":").grid(row=row, column=0, padx=5, pady=5, sticky="e")

            # 创建Entry控件
            entry = ttk.Entry(main_frame, width=30)

            # 如果是编辑模式，填充数据
            if data:
                # data索引从0开始，但字段数据从索引1开始
                value_index = row + 1  # name在data[1], manager在data[2], contact在data[3], address在data[4]
                if value_index < len(data) and data[value_index] is not None:
                    value = str(data[value_index]).strip()
                    entry.insert(0, value)

            # 权限控制：科长只能编辑联系方式和地址
            if current_user['role'] == 'section_chief' and data:
                if field in ["name"]:  # 部门名称和负责人不能修改
                    entry.config(state="readonly")
                elif field in ["manager", "contact", "address"]:  # 联系方式和地址可以编辑
                    entry.config(state="normal")
            elif current_user['role'] == 'staff' and data:
                # 员工只能查看，不能编辑任何字段
                entry.config(state="readonly")

            entry.grid(row=row, column=1, padx=5, pady=5)
            entries[field] = entry

        # 按钮区域
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)

        ttk.Button(btn_frame, text="保存",
                   command=lambda: self.save_department(dialog, entries, data[0] if data else None)).pack(side="left",
                                                                                                          padx=5)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side="left", padx=5)

    def save_department(self, dialog, entries, dept_id=None):
        """保存部门信息"""
        name = entries["name"].get().strip()
        manager = entries["manager"].get().strip()
        contact = entries["contact"].get().strip()
        address = entries["address"].get().strip()

        # 验证必填字段
        if not name or not manager or not address:
            missing_fields = []
            if not name:
                missing_fields.append("部门名称")
            if not manager:
                missing_fields.append("负责人")
            if not address:
                missing_fields.append("地址")

            messagebox.showerror("错误", f"以下字段不能为空:\n{', '.join(missing_fields)}")
            return

        # 获取当前用户
        current_user = self.auth.get_current_user()

        # 如果是科长编辑部门，确保只能编辑自己部门
        if current_user['role'] == 'section_chief' and dept_id:
            user_dept_id = current_user.get('department_id')
            if user_dept_id != dept_id:
                messagebox.showerror("权限不足", "您只能编辑自己所在的部门")
                return
            # 科长不能更改部门名称
            name = entries["name"].get().strip()  # 保持原值不变
        elif current_user['role'] == 'staff':
            messagebox.showerror("权限不足", "员工不能编辑部门信息")
            return

        conn = self.db.get_connection()
        c = conn.cursor()

        try:
            if dept_id:  # 编辑模式
                # 如果是科长，可以更新负责人、联系方式和地址
                if current_user['role'] == 'section_chief':
                    c.execute("UPDATE departments SET manager=?, contact=?, address=? WHERE id=?",
                              (manager, contact, address, dept_id))
                else:
                    # 管理员可以更新所有字段
                    c.execute("UPDATE departments SET name=?, manager=?, contact=?, address=? WHERE id=?",
                              (name, manager, contact, address, dept_id))
            else:
                if current_user['role'] != 'admin':
                    messagebox.showerror("权限不足", "只有管理员可以添加部门")
                    return

                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                c.execute(
                    "INSERT INTO departments (name, manager, contact, address, created_at) VALUES (?, ?, ?, ?, ?)",
                    (name, manager, contact, address, current_time))

            conn.commit()
            messagebox.showinfo("成功", "部门信息已保存")
            dialog.destroy()
            self.load_departments()

        except Exception as e:
            conn.rollback()
            messagebox.showerror("错误", f"保存失败: {str(e)}")
        finally:
            conn.close()

    def delete_department(self):
        """删除部门"""
        selected_item = self.dept_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一条记录")
            return

        # 获取真实的部门ID
        tags = self.dept_tree.item(selected_item, "tags")
        if not tags:
            messagebox.showerror("错误", "无法获取部门ID")
            return

        dept_id = int(tags[0].split("_")[1])
        dept_name = self.dept_tree.item(selected_item)["values"][1]

        # 先检查该部门下是否有员工
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM employees WHERE department_id=?", (dept_id,))
        employee_count = c.fetchone()[0]
        conn.close()

        # 构建确认消息
        if employee_count > 0:
            confirm_message = f"确定要删除部门 '{dept_name}' 吗？\n\n⚠️ 注意：该部门下有 {employee_count} 名员工，删除部门将同时删除这些员工。"
        else:
            confirm_message = f"确定要删除部门 '{dept_name}' 吗？"

        # 单次确认
        if messagebox.askyesno("确认删除", confirm_message):
            conn = self.db.get_connection()
            c = conn.cursor()

            try:
                # 先删除该部门的所有员工
                if employee_count > 0:
                    c.execute("DELETE FROM employees WHERE department_id=?", (dept_id,))

                # 再删除部门
                c.execute("DELETE FROM departments WHERE id=?", (dept_id,))
                conn.commit()

                # 刷新部门列表
                self.load_departments()

                # 显示删除结果
                if employee_count > 0:
                    messagebox.showinfo("删除成功", f"部门 '{dept_name}' 及其 {employee_count} 名员工已删除")
                else:
                    messagebox.showinfo("删除成功", f"部门 '{dept_name}' 已删除")

            except Exception as e:
                conn.rollback()
                messagebox.showerror("删除失败", f"删除操作失败: {e}")
            finally:
                conn.close()

    def manage_employees(self):
        """管理员工对话框，管理员可以管理所有部门，科长只能管理自己部门的员工"""
        selected_item = self.dept_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一个部门")
            return

        # 获取真实的部门ID
        tags = self.dept_tree.item(selected_item, "tags")
        if not tags:
            messagebox.showerror("错误", "无法获取部门ID")
            return

        dept_id = int(tags[0].split("_")[1])
        dept_name = self.dept_tree.item(selected_item)["values"][1]

        # 获取当前用户信息
        current_user = self.auth.get_current_user()

        # 如果是科长，检查是否是自己所在的部门
        if current_user['role'] == 'section_chief':
            user_dept_id = current_user.get('department_id')
            if user_dept_id != dept_id:
                messagebox.showerror("权限不足", "您只能管理自己所在部门的员工")
                return

        # 创建管理员工对话框
        self.employee_dialog_window = tk.Toplevel(self.parent)
        self.employee_dialog_window.title(f"管理员工 - {dept_name}")
        self.employee_dialog_window.geometry("800x500")

        # 设置对话框属性
        self.employee_dialog_window.grab_set()  # 设置为模态对话框
        self.employee_dialog_window.focus_set()  # 获取焦点
        self.employee_dialog_window.transient(self.parent)  # 关联到父窗口

        # 居中显示
        self.utils.center_window(self.employee_dialog_window)

        # 存储当前部门ID和名称
        self.current_dept_id = dept_id
        self.current_dept_name = dept_name

        # 显示当前部门信息
        dept_info_frame = ttk.Frame(self.employee_dialog_window, padding=5)
        dept_info_frame.pack(fill="x")

        ttk.Label(dept_info_frame, text=f"当前部门: {dept_name}",
                  font=('微软雅黑', 12, 'bold')).pack(anchor="w")

        # 工具栏，科长可以添加、编辑、删除本部门员工
        toolbar = ttk.Frame(self.employee_dialog_window, padding=5)
        toolbar.pack(fill="x")

        ttk.Button(toolbar, text="添加员工", command=self.add_employee).pack(side="left", padx=5)
        ttk.Button(toolbar, text="编辑员工", command=self.edit_employee).pack(side="left", padx=5)
        ttk.Button(toolbar, text="删除员工", command=self.delete_employee).pack(side="left", padx=5)
        ttk.Button(toolbar, text="刷新", command=lambda: self.load_employees()).pack(side="left", padx=5)

        # 员工表格
        tree_frame = ttk.Frame(self.employee_dialog_window)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.emp_tree = ttk.Treeview(tree_frame, columns=("序号", "姓名", "职位", "联系方式", "排序号"), show="headings")

        # 设置列
        columns = {
            "序号": {"width": 50, "anchor": "center"},
            "姓名": {"width": 100, "anchor": "center"},
            "职位": {"width": 120, "anchor": "center"},
            "联系方式": {"width": 150, "anchor": "center"},
            "排序号": {"width": 80, "anchor": "center"}
        }

        for col, settings in columns.items():
            self.emp_tree.column(col, **settings)
            self.emp_tree.heading(col, text=col)

        # 绑定双击事件查看员工详情
        self.emp_tree.bind("<Double-1>", self.show_employee_detail)

        # 滚动条
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.emp_tree.yview)
        self.emp_tree.configure(yscrollcommand=vsb.set)

        self.emp_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 加载员工数据
        self.load_employees()

    def load_employees(self):
        """加载员工数据"""
        if not hasattr(self, 'emp_tree') or not hasattr(self, 'current_dept_id'):
            return

        for item in self.emp_tree.get_children():
            self.emp_tree.delete(item)

        conn = self.db.get_connection()
        c = conn.cursor()

        # 先按sort_order排序，再按id排序
        c.execute(
            "SELECT id, name, position, contact, sort_order FROM employees WHERE department_id=? ORDER BY sort_order, id",
            (self.current_dept_id,))
        rows = c.fetchall()
        conn.close()

        # 显示序号和排序号
        for index, row in enumerate(rows, 1):
            # 使用正确的排序号，如果排序号为空则使用显示序号
            sort_order = row[4] if row[4] is not None else index
            display_values = (index, row[1], row[2], row[3], sort_order)
            self.emp_tree.insert("", "end", values=display_values, tags=(f"emp_{row[0]}",))

    def add_employee(self):
        """添加员工对话框"""
        if not hasattr(self, 'current_dept_id') or not hasattr(self, 'current_dept_name'):
            return

        self.employee_dialog("添加员工")

    def edit_employee(self):
        """编辑员工对话框"""
        if not hasattr(self, 'emp_tree'):
            return

        selected_item = self.emp_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一个员工")
            return

        # 获取真实的员工ID
        tags = self.emp_tree.item(selected_item, "tags")
        if not tags:
            messagebox.showerror("错误", "无法获取员工ID")
            return

        emp_id = int(tags[0].split("_")[1])
        emp_name = self.emp_tree.item(selected_item)["values"][1]

        # 从数据库获取真实数据
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT id, name, position, contact, sort_order FROM employees WHERE id=?", (emp_id,))
        real_data = c.fetchone()
        conn.close()

        if not real_data:
            messagebox.showerror("错误", "未找到员工信息")
            return

        self.employee_dialog("编辑员工", real_data)

    def employee_dialog(self, title, data=None):
        """员工添加/编辑对话框"""
        if not hasattr(self, 'current_dept_id') or not hasattr(self, 'current_dept_name'):
            return

        dialog = tk.Toplevel(self.employee_dialog_window)
        dialog.title(title)
        dialog.transient(self.employee_dialog_window)
        dialog.grab_set()
        dialog.geometry("400x350")

        # 居中显示在父对话框上
        self.utils.center_window_on_parent(dialog, self.employee_dialog_window)

        # 主框架
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill="both", expand=True)

        # 显示所属部门信息
        dept_frame = ttk.Frame(main_frame)
        dept_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(dept_frame, text="所属部门:", font=('微软雅黑', 10, 'bold')).pack(side="left")
        ttk.Label(dept_frame, text=self.current_dept_name, font=('微软雅黑', 10)).pack(side="left", padx=(5, 0))

        # 表单框架
        form_frame = ttk.Frame(main_frame)
        form_frame.pack(fill="x", pady=(0, 10))

        entries = {}

        # 表单字段
        fields = [
            ("姓名", "name", 0),
            ("职位", "position", 1),
            ("联系方式", "contact", 2),
            ("排序号", "sort_order", 3)
        ]

        for label, field, row in fields:
            ttk.Label(form_frame, text=label + ":").grid(row=row, column=0, padx=5, pady=5, sticky="e")

            if field == "sort_order":
                # 排序号使用Spinbox
                entry = ttk.Spinbox(form_frame, from_=1, to=999, width=27)
            else:
                entry = ttk.Entry(form_frame, width=30)

            entry.grid(row=row, column=1, padx=5, pady=5)
            entries[field] = entry

        # 如果是编辑模式填充数据
        if data:
            entries["name"].insert(0, data[1])
            entries["position"].insert(0, data[2] if len(data) > 2 else "")
            entries["contact"].insert(0, data[3] if len(data) > 3 else "")
            # 设置排序号
            sort_order_value = data[4] if len(data) > 4 and data[4] is not None else 1
            entries["sort_order"].delete(0, tk.END)
            entries["sort_order"].insert(0, str(sort_order_value))

        # 按钮区域
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text="保存",
                   command=lambda: self.save_employee(dialog, entries, data[0] if data else None)).pack(side="left",
                                                                                                        padx=5)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side="left", padx=5)

    def save_employee(self, dialog, entries, emp_id=None):
        """保存员工信息"""
        name = entries["name"].get().strip()
        position = entries["position"].get().strip()
        contact = entries["contact"].get().strip()

        # 获取排序号
        try:
            sort_order = int(entries["sort_order"].get())
        except ValueError:
            sort_order = 1  # 默认值为1

        # 验证必填字段
        if not name:
            messagebox.showerror("错误", "员工姓名不能为空")
            return

        conn = self.db.get_connection()
        c = conn.cursor()

        try:
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if emp_id:  # 编辑模式
                c.execute("UPDATE employees SET name=?, position=?, contact=?, sort_order=? WHERE id=?",
                          (name, position, contact, sort_order, emp_id))
            else:  # 新增模式
                c.execute(
                    "INSERT INTO employees (name, position, department_id, contact, sort_order, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (name, position, self.current_dept_id, contact, sort_order, current_time))

            conn.commit()
            messagebox.showinfo("成功", "员工信息已保存")
            dialog.destroy()

            # 刷新员工列表
            self.load_employees()

            # 确保主对话框保持显示状态
            if hasattr(self, 'employee_dialog_window') and self.employee_dialog_window.winfo_exists():
                self.employee_dialog_window.lift()
                self.employee_dialog_window.focus_force()
                self.employee_dialog_window.deiconify()

        except Exception as e:
            conn.rollback()
            messagebox.showerror("错误", f"保存失败: {str(e)}")
        finally:
            conn.close()

    def delete_employee(self):
        """删除员工"""
        if not hasattr(self, 'emp_tree'):
            return

        selected_item = self.emp_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一个员工")
            return

        # 获取真实的员工ID
        tags = self.emp_tree.item(selected_item, "tags")
        if not tags:
            messagebox.showerror("错误", "无法获取员工ID")
            return

        emp_id = int(tags[0].split("_")[1])
        emp_name = self.emp_tree.item(selected_item)["values"][1]

        if messagebox.askyesno("确认", f"确定要删除员工 '{emp_name}' 吗？"):
            conn = self.db.get_connection()
            c = conn.cursor()

            try:
                c.execute("DELETE FROM employees WHERE id=?", (emp_id,))
                conn.commit()

                # 刷新员工列表
                self.load_employees()

                # 确保对话框保持显示状态
                if hasattr(self, 'employee_dialog_window') and self.employee_dialog_window.winfo_exists():
                    # 将对话框提到最前面
                    self.employee_dialog_window.lift()
                    self.employee_dialog_window.focus_force()
                    # 确保窗口显示
                    self.employee_dialog_window.deiconify()

                messagebox.showinfo("成功", f"员工 '{emp_name}' 已删除")

            except Exception as e:
                conn.rollback()
                messagebox.showerror("错误", f"删除失败: {e}")
            finally:
                conn.close()

    def show_employee_detail(self, event):
        """显示员工详情"""
        selected_item = self.emp_tree.selection()
        if not selected_item:
            return

        # 获取真实的员工ID
        tags = self.emp_tree.item(selected_item, "tags")
        if not tags:
            return

        emp_id = int(tags[0].split("_")[1])

        # 从数据库获取完整信息
        conn = self.db.get_connection()
        c = conn.cursor()
        c.execute("SELECT id, name, position, contact, sort_order FROM employees WHERE id=?", (emp_id,))
        emp_data = c.fetchone()
        conn.close()

        if not emp_data:
            return

        # 获取部门名称
        dept_name = self.current_dept_name if hasattr(self, 'current_dept_name') else "未知部门"

        # 创建详情对话框
        detail_dialog = tk.Toplevel(self.employee_dialog_window)
        detail_dialog.title(f"员工详情 - {emp_data[1]}")
        detail_dialog.geometry("400x300")

        # 设置对话框属性
        detail_dialog.transient(self.employee_dialog_window)
        detail_dialog.grab_set()
        detail_dialog.focus_set()

        # 居中显示在父窗口上
        self.utils.center_window_on_parent(detail_dialog, self.employee_dialog_window)

        # 主框架
        main_frame = ttk.Frame(detail_dialog, padding=20)
        main_frame.pack(fill="both", expand=True)

        # 员工基本信息
        info_frame = ttk.LabelFrame(main_frame, text="员工信息", padding=10)
        info_frame.pack(fill="x", pady=(0, 10))

        info_text = f"""
    员工ID：{emp_data[0]}
    姓  名：{emp_data[1]}
    所属部门：{dept_name}
    职  位：{emp_data[2] or '未设置'}
    联系方式：{emp_data[3] or '未设置'}
    排序号：{emp_data[4] if emp_data[4] is not None else '未设置'}
        """

        info_label = tk.Label(info_frame, text=info_text, justify="left", font=('微软雅黑', 11))
        info_label.pack(anchor="w")

        # 关闭按钮
        def close_detail():
            detail_dialog.grab_release()
            detail_dialog.destroy()

        ttk.Button(main_frame, text="关闭", command=close_detail).pack(pady=10)

        # 绑定ESC键关闭
        detail_dialog.bind('<Escape>', lambda e: close_detail())

    def export_departments_to_excel(self):
        """导出部门数据到Excel"""
        try:
            if not self.auth.has_permission("departments", "export"):
                messagebox.showerror("权限不足", "您没有导出权限")
                return

            default_filename = f"部门员工列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialfile=default_filename
            )

            if not filepath:
                return

            conn = self.db.get_connection()
            c = conn.cursor()

            # 按照ID排序，与界面显示顺序一致
            c.execute("SELECT id, name, manager, contact, address FROM departments ORDER BY id")
            dept_rows = c.fetchall()

            if not dept_rows:
                messagebox.showwarning("提示", "没有找到部门记录")
                return

            # 创建Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "部门员工清单"

            # 添加标题行，与导出模板格式一致
            headers = ["部门名称", "负责人", "联系方式", "部门地址",
                       "部门内员工序号", "员工姓名", "员工职位", "员工联系方式"]
            ws.append(headers)

            # 遍历每个部门，导出部门信息和员工信息
            total_depts = 0
            total_employees = 0
            unique_employees = set()

            for dept_row in dept_rows:
                dept_id, dept_name, dept_manager, dept_contact, dept_address = dept_row
                total_depts += 1

                # 检查负责人是否是有效的员工
                if dept_manager:
                    unique_employees.add(dept_manager)

                # 按sort_order排序
                c.execute("SELECT name, position, contact FROM employees WHERE department_id=? ORDER BY sort_order, id",
                          (dept_id,))
                employees = c.fetchall()

                # 部门内员工序号
                emp_seq = 1

                if employees:
                    for i, emp in enumerate(employees, 1):
                        emp_name, emp_position, emp_contact = emp

                        # 添加员工到集合
                        unique_employees.add(emp_name)

                        if i == 1:
                            ws.append([
                                dept_name, dept_manager, dept_contact, dept_address,
                                f"{emp_seq:02d}", emp_name, emp_position, emp_contact
                            ])
                        else:
                            ws.append([
                                "", "", "", "",
                                f"{emp_seq:02d}", emp_name, emp_position, emp_contact
                            ])
                        emp_seq += 1
                else:
                    ws.append([
                        dept_name, dept_manager, dept_contact, dept_address,
                        "", "", "", ""
                    ])

            conn.close()

            # 更新总员工数为去重后的员工数量
            total_employees = len(unique_employees)

            # 添加统计信息
            ws.append([])
            ws.append(["统计信息"])
            ws.append(["总部门数", total_depts])
            ws.append(["总员工数", total_employees])

            conn.close()

            ws.append([])
            # 将多行说明合并到一行中，使用换行符分隔
            export_note = "导出说明：\n" \
                          "1. 同一部门的员工需要连续填写，部门信息只在第一行填写。\n" \
                          "2. 部门内员工序号从01开始连续编号。\n" \
                          "3. 如果只有部门没有员工，员工信息列可以留空。\n" \
                          "4. 导入时会根据部门名称自动分组处理。\n" \
                          "5. 请勿修改标题行的列名和顺序。"
            ws.append([export_note])

            ws.append([])
            ws.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

            # 设置列宽
            col_widths = [20, 15, 15, 50, 15, 15, 15, 15]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            # 设置标题样式
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True, name='微软雅黑', size=11)
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")

            # 设置数据区域边框
            from openpyxl.styles import Border, Side
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # 计算数据区域行数
            data_row_count = 1
            for dept_row in dept_rows:
                dept_id = dept_row[0]
                temp_conn = self.db.get_connection()
                temp_c = temp_conn.cursor()
                temp_c.execute("SELECT COUNT(*) FROM employees WHERE department_id=?", (dept_id,))
                emp_count = temp_c.fetchone()[0]
                temp_conn.close()

                data_row_count += max(1, emp_count)

            # 为数据区域设置边框
            for row in ws.iter_rows(min_row=1, max_row=data_row_count, max_col=8):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

            # 保存文件
            wb.save(filepath)
            messagebox.showinfo("导出成功",
                                f"成功导出数据到:\n{filepath}\n\n"
                                f"• 部门数量: {total_depts}\n"
                                f"• 员工数量: {total_employees}（含负责人）")

        except PermissionError:
            messagebox.showerror("错误", "无文件写入权限，请选择其他位置")
        except Exception as e:
            messagebox.showerror("错误", f"导出过程中出错: {str(e)}")

    def import_departments_from_excel(self):
        """从Excel导入部门数据"""
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")],
            title="选择要导入的Excel文件"
        )

        if not filepath:
            return

        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            # 获取标题行
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

            print(f"检测到的标题行: {headers}")

            # 判断是哪种格式
            if headers == ["部门名称", "负责人", "联系方式", "部门地址",
                           "部门内员工序号", "员工姓名", "员工职位", "员工联系方式"]:
                return self._import_template_format(wb, headers)
            elif headers == ["部门名称", "部门地址", "部门联系方式", "部门负责人",
                             "部门内员工序号", "员工姓名", "员工职位", "员工联系方式"]:
                return self._import_old_template_format(wb, headers)
            elif headers == ["部门ID", "部门名称", "负责人", "联系方式", "地址"]:
                return self._import_department_list_format(wb, headers)
            elif len(headers) >= 5 and headers[0] == "ID" and "部门名称" in headers[1]:
                return self._import_department_list_format(wb, headers)
            else:
                return self._try_smart_import(wb, headers)

        except Exception as e:
            messagebox.showerror("错误", f"导入失败: {str(e)}")
            import traceback
            traceback.print_exc()

    def _import_template_format(self, wb, headers):
        """导入模板格式，格式：部门名称、负责人、联系方式、部门地址等"""
        ws = wb.active

        conn = self.db.get_connection()
        c = conn.cursor()
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        dept_success_count = 0
        emp_success_count = 0
        current_dept_id = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            (dept_name, dept_manager, dept_contact, dept_address,
             emp_seq, emp_name, emp_position, emp_contact) = row

            # 清理数据
            dept_name = str(dept_name).strip() if dept_name else ""
            dept_manager = str(dept_manager).strip() if dept_manager else ""
            dept_contact = str(dept_contact).strip() if dept_contact else ""
            dept_address = str(dept_address).strip() if dept_address else ""
            emp_name = str(emp_name).strip() if emp_name else ""
            emp_position = str(emp_position).strip() if emp_position else ""
            emp_contact = str(emp_contact).strip() if emp_contact else ""

            # 如果有部门信息
            if dept_name:
                c.execute('''SELECT id FROM departments 
                           WHERE name=? AND manager=? AND address=?''',
                          (dept_name, dept_manager, dept_address))
                existing = c.fetchone()

                if existing:
                    current_dept_id = existing[0]
                else:
                    c.execute('''INSERT INTO departments 
                               (name, manager, contact, address, created_at) 
                               VALUES (?, ?, ?, ?, ?)''',
                              (dept_name, dept_manager, dept_contact, dept_address, current_time))
                    current_dept_id = c.lastrowid
                    dept_success_count += 1

            # 如果有员工信息且部门已确定
            if emp_name and current_dept_id:
                c.execute("SELECT id FROM employees WHERE name=? AND department_id=?",
                          (emp_name, current_dept_id))
                existing = c.fetchone()

                if not existing:
                    # 获取排序号
                    sort_order = 1
                    if emp_seq:
                        try:
                            sort_order = int(str(emp_seq).strip())
                        except ValueError:
                            sort_order = 1

                    c.execute('''INSERT INTO employees 
                               (name, position, department_id, contact, sort_order, created_at) 
                               VALUES (?, ?, ?, ?, ?, ?)''',
                              (emp_name, emp_position, current_dept_id, emp_contact, sort_order, current_time))
                    emp_success_count += 1

        conn.commit()
        conn.close()

        messagebox.showinfo("导入成功",
                            f"导入完成！\n\n"
                            f"新增部门：{dept_success_count} 个\n"
                            f"新增员工：{emp_success_count} 名")

        self.load_departments()

    def _import_old_template_format(self, wb, headers):
        """导入旧的模板格式"""
        ws = wb.active

        conn = self.db.get_connection()
        c = conn.cursor()
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        dept_success_count = 0
        emp_success_count = 0
        current_dept_id = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            (dept_name, dept_address, dept_contact, dept_manager,
             emp_seq, emp_name, emp_position, emp_contact) = row

            # 清理数据
            dept_name = str(dept_name).strip() if dept_name else ""
            dept_address = str(dept_address).strip() if dept_address else ""
            dept_contact = str(dept_contact).strip() if dept_contact else ""
            dept_manager = str(dept_manager).strip() if dept_manager else ""
            emp_name = str(emp_name).strip() if emp_name else ""
            emp_position = str(emp_position).strip() if emp_position else ""
            emp_contact = str(emp_contact).strip() if emp_contact else ""

            # 如果有部门信息
            if dept_name:
                c.execute("SELECT id FROM departments WHERE name=?", (dept_name,))
                existing = c.fetchone()

                if existing:
                    current_dept_id = existing[0]
                else:
                    c.execute('''INSERT INTO departments 
                               (name, manager, contact, address, created_at) 
                               VALUES (?, ?, ?, ?, ?)''',
                              (dept_name, dept_manager, dept_contact, dept_address, current_time))
                    current_dept_id = c.lastrowid
                    dept_success_count += 1

            # 如果有员工信息且部门已确定
            if emp_name and current_dept_id:
                c.execute("SELECT id FROM employees WHERE name=? AND department_id=?",
                          (emp_name, current_dept_id))
                existing = c.fetchone()

                if not existing:
                    # 获取排序号
                    sort_order = 1
                    if emp_seq:
                        try:
                            sort_order = int(str(emp_seq).strip())
                        except ValueError:
                            sort_order = 1

                    c.execute('''INSERT INTO employees 
                               (name, position, department_id, contact, sort_order, created_at) 
                               VALUES (?, ?, ?, ?, ?, ?)''',
                              (emp_name, emp_position, current_dept_id, emp_contact, sort_order, current_time))
                    emp_success_count += 1

        conn.commit()
        conn.close()

        messagebox.showinfo("导入成功",
                            f"导入完成！\n\n"
                            f"新增部门：{dept_success_count} 个\n"
                            f"新增员工：{emp_success_count} 名")

        self.load_departments()

    def _import_department_list_format(self, wb, headers):
        """导入部门列表格式"""
        ws = wb.active

        conn = self.db.get_connection()
        c = conn.cursor()
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        success_count = 0
        duplicate_count = 0

        # 确定列索引
        col_mapping = {}
        for idx, header in enumerate(headers):
            header_str = str(header).strip().lower()
            if "id" in header_str:
                col_mapping['id'] = idx
            elif "部门名称" in header_str or "名称" in header_str:
                col_mapping['name'] = idx
            elif "负责人" in header_str:
                col_mapping['manager'] = idx
            elif "联系方式" in header_str or "联系" in header_str:
                col_mapping['contact'] = idx
            elif "地址" in header_str:
                col_mapping['address'] = idx

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            dept_name = str(row[col_mapping.get('name', 1)]).strip() if len(row) > col_mapping.get('name', 1) else ""
            dept_manager = str(row[col_mapping.get('manager', 2)]).strip() if len(row) > col_mapping.get('manager',
                                                                                                         2) else ""
            dept_contact = str(row[col_mapping.get('contact', 3)]).strip() if len(row) > col_mapping.get('contact',
                                                                                                         3) else ""
            dept_address = str(row[col_mapping.get('address', 4)]).strip() if len(row) > col_mapping.get('address',
                                                                                                         4) else ""

            if not dept_name:
                continue

            # 检查是否已存在
            c.execute("SELECT id FROM departments WHERE name=?", (dept_name,))
            existing = c.fetchone()

            if existing:
                duplicate_count += 1
            else:
                c.execute('''INSERT INTO departments 
                           (name, manager, contact, address, created_at) 
                           VALUES (?, ?, ?, ?, ?)''',
                          (dept_name, dept_manager, dept_contact, dept_address, current_time))
                success_count += 1

        conn.commit()
        conn.close()

        messagebox.showinfo("导入结果",
                            f"导入完成！\n\n"
                            f"新增部门：{success_count} 个\n"
                            f"重复跳过：{duplicate_count} 个")

        self.load_departments()

    def _try_smart_import(self, wb, headers):
        """尝试智能识别导入格式"""
        # 检查是否是纯部门数据
        if len(headers) >= 4:
            fields_found = 0
            for header in headers:
                header_str = str(header).strip().lower()
                if any(keyword in header_str for keyword in ["部门", "名称", "name"]):
                    fields_found += 1
                elif any(keyword in header_str for keyword in ["负责", "经理", "manager"]):
                    fields_found += 1
                elif any(keyword in header_str for keyword in ["联系", "电话", "contact", "phone"]):
                    fields_found += 1
                elif any(keyword in header_str for keyword in ["地址", "地点", "address", "location"]):
                    fields_found += 1

            if fields_found >= 3:
                return self._import_department_list_format(wb, headers)

        # 检查是否是部门+员工合并格式
        if len(headers) >= 8:
            emp_fields_found = 0
            for header in headers:
                header_str = str(header).strip().lower()
                if any(keyword in header_str for keyword in ["员工", "姓名", "职员", "employee", "staff"]):
                    emp_fields_found += 1
                elif any(keyword in header_str for keyword in ["职位", "岗位", "position", "job"]):
                    emp_fields_found += 1

            if emp_fields_found >= 2:
                return self._import_template_format(wb, headers)

        messagebox.showerror("错误",
                             f"无法识别的Excel格式\n\n"
                             f"检测到的标题行：\n{headers}\n\n"
                             f"请使用以下格式之一：\n"
                             f"1. 部门模板格式（含员工）\n"
                             f"2. 纯部门列表格式\n"
                             f"3. 系统导出的标准格式")

    def export_template(self):
        """导出导入模板"""
        try:
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialfile="部门员工导入模板.xlsx"
            )

            if not filepath:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "部门员工清单"

            # 与export_departments_to_excel使用相同的标题行
            headers = ["部门名称", "负责人", "联系方式", "部门地址",
                       "部门内员工序号", "员工姓名", "员工职位", "员工联系方式"]
            ws.append(headers)

            # 与export_departments_to_excel使用相同的数据格式
            example_data = [
                ["技术部", "张经理", "010-12345678", "北京市海淀区中关村科技园区A座101室", "01", "张三", "工程师",
                 "13800138000"],
                ["", "", "", "", "02", "李四", "设计师", "13900139000"],
                ["销售部", "王总监", "021-87654321", "上海市浦东新区陆家嘴金融中心B座202室", "01", "王五", "销售经理",
                 "13700137000"],
                ["", "", "", "", "02", "赵六", "销售代表", "13600136000"],
                ["财务部", "刘会计", "020-55555555", "广州市天河区珠江新城C座303室", "", "", "", ""]
            ]

            for row in example_data:
                ws.append(row)

            # 与export_departments_to_excel使用相同的列宽
            col_widths = [20, 15, 15, 50, 15, 15, 15, 15]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            # 设置标题样式
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True, name='微软雅黑', size=11)
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

            # 设置数据区域边框
            from openpyxl.styles import Border, Side
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # 为数据区域设置边框
            for row in ws.iter_rows(min_row=1, max_row=len(example_data) + 1, max_col=8):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

            # 添加说明信息，格式与export_departments_to_excel保持一致
            ws.append([])

            # 与export_departments_to_excel使用相同的说明格式
            export_note = "导出说明：\n" \
                          "1. 同一部门的员工需要连续填写，部门信息只在第一行填写。\n" \
                          "2. 部门内员工序号从01开始连续编号。\n" \
                          "3. 如果只有部门没有员工，员工信息列可以留空。\n" \
                          "4. 导入时会根据部门名称自动分组处理。\n" \
                          "5. 请勿修改标题行的列名和顺序。"
            ws.append([export_note])

            ws.append([])
            ws.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

            # 设置说明行的换行格式
            from openpyxl.styles import Alignment
            note_row = ws.max_row - 2  # 说明行
            note_cell = ws.cell(row=note_row, column=1)
            note_cell.alignment = Alignment(wrapText=True, vertical="top")

            # 调整说明行高度
            ws.row_dimensions[note_row].height = 80

            wb.save(filepath)
            messagebox.showinfo("导出成功", f"模板已导出到:\n{filepath}")

        except Exception as e:
            messagebox.showerror("错误", f"导出模板失败: {str(e)}")

    def show_department_detail(self, event):
        """显示部门详情"""
        selected_item = self.dept_tree.selection()
        if not selected_item:
            return

        # 获取真实的部门ID
        tags = self.dept_tree.item(selected_item, "tags")
        if not tags:
            messagebox.showerror("错误", "无法获取部门ID")
            return

        dept_id = int(tags[0].split("_")[1])

        conn = self.db.get_connection()
        c = conn.cursor()

        c.execute('''SELECT d.*, COUNT(e.id) as employee_count 
                     FROM departments d 
                     LEFT JOIN employees e ON d.id = e.department_id 
                     WHERE d.id=? 
                     GROUP BY d.id''', (dept_id,))
        dept_data = c.fetchone()

        if not dept_data:
            messagebox.showerror("错误", "未找到部门信息")
            conn.close()
            return

        # 获取部门员工列表，按sort_order排序
        c.execute("SELECT name, position, contact FROM employees WHERE department_id=? ORDER BY sort_order, id",
                  (dept_id,))
        employees = c.fetchall()

        conn.close()

        detail_dialog = tk.Toplevel(self.parent)
        detail_dialog.title(f"部门详情 - {dept_data[1]}")
        detail_dialog.geometry("600x500")

        self.utils.center_window(detail_dialog)

        main_frame = ttk.Frame(detail_dialog, padding=20)
        main_frame.pack(fill="both", expand=True)

        info_frame = ttk.LabelFrame(main_frame, text="部门信息", padding=10)
        info_frame.pack(fill="x", pady=(0, 10))

        info_text = f"""
    部门名称：{dept_data[1]}
    负 责 人：{dept_data[2] or '未设置'}
    联系方式：{dept_data[3] or '未设置'}
    地    址：{dept_data[4]}
    创建时间：{dept_data[5]}
    员工数量：{dept_data[6]} 人
    """

        info_label = tk.Label(info_frame, text=info_text, justify="left", font=('微软雅黑', 11))
        info_label.pack(anchor="w")

        if employees:
            emp_frame = ttk.LabelFrame(main_frame, text="员工列表", padding=10)
            emp_frame.pack(fill="both", expand=True)

            tree_frame = ttk.Frame(emp_frame)
            tree_frame.pack(fill="both", expand=True)

            emp_tree = ttk.Treeview(tree_frame, columns=("序号", "姓名", "职位", "联系方式"), show="headings", height=8)

            columns = {
                "序号": {"width": 50, "anchor": "center"},
                "姓名": {"width": 100, "anchor": "center"},
                "职位": {"width": 120, "anchor": "center"},
                "联系方式": {"width": 150, "anchor": "center"}
            }

            for col, settings in columns.items():
                emp_tree.column(col, **settings)
                emp_tree.heading(col, text=col)

            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=emp_tree.yview)
            emp_tree.configure(yscrollcommand=vsb.set)

            emp_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")

            tree_frame.grid_rowconfigure(0, weight=1)
            tree_frame.grid_columnconfigure(0, weight=1)

            for index, emp in enumerate(employees, 1):
                emp_tree.insert("", "end", values=(index, emp[0], emp[1], emp[2]))
        else:
            no_emp_label = tk.Label(main_frame, text="该部门暂无员工", font=('微软雅黑', 11), fg="gray")
            no_emp_label.pack(pady=10)

        ttk.Button(main_frame, text="关闭", command=detail_dialog.destroy).pack(pady=10)
